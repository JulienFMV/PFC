#!/usr/bin/env python3
"""Export Swiss CT hourly PFC to Analyse_HPFC.xlsx."""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_PFC = ROOT / "pfc_shaping" / "output" / f"pfc_15min_{date.today().isoformat()}.parquet"
DEFAULT_WORKBOOK = Path(r"H:\Energy\GeCom\SYSTEM\TS-Energy\04 Analyses\9 - HPFC\Analyse_HPFC.xlsx")
P4_SHEET = "P4 - JB"
PARAM_SHEET = "Paramètres "


def _hourly_from_pfc(path: Path) -> pd.DataFrame:
    pfc = pd.read_parquet(path)
    hourly = (
        pfc["price_shape"]
        .tz_convert("Europe/Zurich")
        .resample("h")
        .mean()
        .to_frame("HPFC P4")
        .reset_index()
        .rename(columns={"index": "Date"})
    )
    hourly["Date"] = hourly["Date"].dt.tz_localize(None)
    return hourly


def _filter_window(hourly: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date) + pd.Timedelta(hours=23)
    out = hourly[(hourly["Date"] >= start_ts) & (hourly["Date"] <= end_ts)].copy()
    if out.empty:
        raise ValueError(f"No hourly CT rows found between {start_date} and {end_date}.")
    return out


def _backup_workbook(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}_backup_{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def _clear_sheet_values(ws, min_row: int = 2, cols: tuple[int, ...] = (1, 2)) -> None:
    for row in range(min_row, ws.max_row + 1):
        for col in cols:
            ws.cell(row=row, column=col).value = None


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Swiss CT PFC to Analyse_HPFC.xlsx")
    parser.add_argument("--pfc", type=Path, default=DEFAULT_PFC)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--start-date", type=str, default=(date.today() + timedelta(days=1)).isoformat())
    parser.add_argument("--end-date", type=str, default=(date.today() + timedelta(days=5)).isoformat())
    args = parser.parse_args()

    start_date = date.fromisoformat(args.start_date)
    end_date = date.fromisoformat(args.end_date)
    if end_date < start_date:
        raise ValueError("end-date must be >= start-date")
    if not args.pfc.exists():
        raise FileNotFoundError(args.pfc)
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    hourly = _hourly_from_pfc(args.pfc)
    export_df = _filter_window(hourly, start_date, end_date)

    backup = _backup_workbook(args.workbook)
    wb = load_workbook(args.workbook)
    ws_param = wb[PARAM_SHEET]
    ws_p4 = wb[P4_SHEET]

    ws_param["B4"] = datetime.combine(start_date, datetime.min.time())
    ws_param["B5"] = datetime.combine(end_date, datetime.min.time())

    _clear_sheet_values(ws_p4, min_row=2, cols=(1, 2))
    for i, row in enumerate(export_df.itertuples(index=False, name=None), start=2):
        ws_p4.cell(row=i, column=1).value = row[0]
        ws_p4.cell(row=i, column=2).value = float(row[1])

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(args.workbook)

    print(
        f"exported_rows={len(export_df)} "
        f"start={export_df['Date'].min()} end={export_df['Date'].max()} "
        f"workbook={args.workbook} backup={backup}"
    )


if __name__ == "__main__":
    main()
