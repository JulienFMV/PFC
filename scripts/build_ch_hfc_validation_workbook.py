"""Build an Excel validation workbook for a CH hourly HFC/PFC CSV."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.audit_ch_pfc_hourly_shape import _load_csv, audit  # noqa: E402
from scripts.audit_ch_hfc_seasonal_coherence import audit as seasonal_audit  # noqa: E402
from scripts.export_local_test_ch_hourly_csv import _calibration_product, _latest_eex_base_prices  # noqa: E402


DEFAULT_CSV = Path("output/ch_hfc_hourly_20260616_20301231_v5_negative_capture.csv")
DEFAULT_OUTPUT = Path("output/ch_hfc_hourly_20260616_20301231_validation.xlsx")
DEFAULT_REPORT = Path(".planning/phases/13-lt-electrification-scenario-shape/CH-HFC-HOURLY-VALIDATION-WORKBOOK.md")
PRICE_COLUMNS = [
    "price_slow_eur_mwh",
    "price_central_eur_mwh",
    "price_fast_eur_mwh",
    "price_weighted_mean_eur_mwh",
    "structural_p10_eur_mwh",
    "structural_p50_eur_mwh",
    "structural_p90_eur_mwh",
    "structural_width_eur_mwh",
]


def _season(month: int) -> str:
    if month in (12, 1, 2):
        return "Winter"
    if month in (3, 4, 5):
        return "Spring"
    if month in (6, 7, 8):
        return "Summer"
    return "Autumn"


def _prepare(csv_path: Path, forwards_path: Path) -> tuple[pd.DataFrame, dict[str, float], dict[str, float]]:
    df = _load_csv(csv_path)
    latest, forwards = _latest_eex_base_prices(forwards_path, market="CH")
    df["date_ch"] = df["ts"].dt.strftime("%d.%m.%Y")
    df["year"] = df["ts"].dt.year.astype(int)
    df["quarter"] = df["ts"].dt.quarter.astype(int)
    df["month"] = df["ts"].dt.month.astype(int)
    df["month_name"] = df["ts"].dt.strftime("%m")
    df["hour"] = df["ts"].dt.hour.astype(int)
    df["weekday"] = df["ts"].dt.weekday.astype(int)
    df["day_type"] = df["weekday"].map(lambda x: "Weekend" if x >= 5 else "Weekday")
    df["season"] = df["month"].map(_season)
    df["calibration_product"] = [_calibration_product(ts, forwards) for ts in df["ts"]]
    df["negative_weighted"] = df["price_weighted_mean_eur_mwh"] < 0.0
    df["negative_p10"] = df["structural_p10_eur_mwh"] < 0.0
    df["very_low_weighted_lt_5"] = df["price_weighted_mean_eur_mwh"] < 5.0
    df["very_low_p10_lt_0"] = df["structural_p10_eur_mwh"] < 0.0
    df["timestamp_ch_excel"] = df["timestamp_ch"]
    df["timestamp_utc_excel"] = df["timestamp_utc"]
    metrics = {
        "latest_eex_ch_base_date": str(pd.Timestamp(latest).date()),
    }
    return df, forwards, metrics


def _eex_controls(df: pd.DataFrame, forwards: dict[str, float]) -> pd.DataFrame:
    rows = []
    for product, group in df.groupby("calibration_product", dropna=False):
        if not product or product not in forwards:
            continue
        row: dict[str, object] = {
            "calibration_product": product,
            "target_eex_base_eur_mwh": float(forwards[product]),
            "rows": len(group),
        }
        for col in PRICE_COLUMNS:
            row[f"mean_{col}"] = float(group[col].mean())
        row["abs_error_weighted_eur_mwh"] = abs(row["mean_price_weighted_mean_eur_mwh"] - row["target_eex_base_eur_mwh"])
        rows.append(row)
    return pd.DataFrame(rows).sort_values("calibration_product").reset_index(drop=True)


def _period_controls(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = ["year", "quarter", "month"]
    return df.groupby(group_cols, as_index=False)[PRICE_COLUMNS].mean().round(6)


def _duck_month(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = ["year", "month", "hour"]
    return df.groupby(group_cols, as_index=False)[PRICE_COLUMNS].mean().round(6)


def _duck_season(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = ["year", "season", "hour"]
    return df.groupby(group_cols, as_index=False)[PRICE_COLUMNS].mean().round(6)


def _heatmap_month_hour(df: pd.DataFrame, *, value_col: str = "price_weighted_mean_eur_mwh") -> pd.DataFrame:
    heat = df.pivot_table(index=["year", "month"], columns="hour", values=value_col, aggfunc="mean").reset_index()
    heat.columns = [str(c) for c in heat.columns]
    return heat.round(6)


def _structural_width(df: pd.DataFrame) -> pd.DataFrame:
    return (
        df.groupby(["year", "month", "hour"], as_index=False)
        .agg(
            structural_width_mean_eur_mwh=("structural_width_eur_mwh", "mean"),
            structural_width_p95_eur_mwh=("structural_width_eur_mwh", lambda x: x.quantile(0.95)),
            structural_p10_mean_eur_mwh=("structural_p10_eur_mwh", "mean"),
            structural_p90_mean_eur_mwh=("structural_p90_eur_mwh", "mean"),
        )
        .round(6)
    )


def _negative_hours(df: pd.DataFrame) -> pd.DataFrame:
    return (
        df.groupby(["year", "month", "hour"], as_index=False)
        .agg(
            rows=("timestamp_ch", "count"),
            weighted_negative_hours=("negative_weighted", "sum"),
            p10_negative_hours=("negative_p10", "sum"),
            weighted_lt_5_hours=("very_low_weighted_lt_5", "sum"),
            p10_lt_0_hours=("very_low_p10_lt_0", "sum"),
            min_weighted_eur_mwh=("price_weighted_mean_eur_mwh", "min"),
            min_p10_eur_mwh=("structural_p10_eur_mwh", "min"),
            min_fast_eur_mwh=("price_fast_eur_mwh", "min"),
        )
        .round(6)
    )


def _chart_data(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    years = sorted(df["year"].unique())
    chart_year = 2030 if 2030 in years else int(years[-1])
    year_df = df[df["year"] == chart_year].copy()
    monthly = year_df.pivot_table(index="hour", columns="month", values="price_weighted_mean_eur_mwh", aggfunc="mean")
    monthly = monthly.reindex(range(24)).round(6).reset_index()
    monthly.columns = ["hour"] + [f"M{int(c):02d}" for c in monthly.columns[1:]]
    seasonal = year_df.pivot_table(index="hour", columns="season", values="price_weighted_mean_eur_mwh", aggfunc="mean")
    seasonal = seasonal.reindex(range(24)).round(6).reset_index()
    width_month = year_df.groupby("month", as_index=False)["structural_width_eur_mwh"].mean().round(6)
    negative_month = (
        year_df.groupby("month", as_index=False)
        .agg(p10_negative_hours=("negative_p10", "sum"), weighted_negative_hours=("negative_weighted", "sum"))
    )
    return {
        "chart_monthly_duck": monthly,
        "chart_seasonal_duck": seasonal,
        "chart_width_month": width_month,
        "chart_negative_month": negative_month,
    }


def _write_report(path: Path, *, workbook: Path, csv_path: Path, metrics: dict[str, float]) -> None:
    lines = [
        "# CH HFC Hourly Validation Workbook",
        "",
        f"* workbook: `{workbook}`",
        f"* source CSV: `{csv_path}`",
        f"* latest EEX CH BASE date: `{metrics['latest_eex_ch_base_date']}`",
        "* scope: `local/test validation`",
        "* production approval: `NO`",
        "",
        "## Workbook Blocks",
        "",
        "| sheet | role |",
        "|---|---|",
        "| `Raw` | full hourly CSV with calendar fields and calibration product |",
        "| `EEX_Means` | EEX product mean control for every price series |",
        "| `Period_Means` | year > quarter > month average price controls |",
        "| `Duck_Month` | hour x month duck curve data by year |",
        "| `Duck_Season` | hour x season duck curve data by year |",
        "| `Heatmap_Month_Hour` | month x hour weighted mean heatmap |",
        "| `Structural_Width` | structural width by year/month/hour |",
        "| `Negative_Low_Hours` | negative and very-low hour diagnostics |",
        "| `Seasonal_Coherence` | January/October and winter/autumn consistency flags |",
        "| `Quoted_EEX_Products` | residuals against every quoted overlapping EEX product |",
        "| `Charts` | prebuilt charts for quick review |",
        "",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def _style_workbook(
    xlsx: Path,
    *,
    include_tables: bool = True,
    include_charts: bool = True,
    hide_chart_data: bool = True,
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        date_cols = {
            cell.column
            for cell in ws[1]
            if cell.value in {"timestamp_ch", "timestamp_utc"}
        }
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for col_idx in date_cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "DD.MM.YYYY HH:MM"
        ws.auto_filter.ref = ws.dimensions
        for idx, col in enumerate(ws.columns, start=1):
            width = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2, 28)
            ws.column_dimensions[get_column_letter(idx)].width = width
        if include_tables and ws.max_row > 1 and ws.max_column > 1 and ws.title not in {"README", "Charts", "Chart_Data"}:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tab = Table(displayName=("T_" + ws.title.replace(" ", "_"))[:255], ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)

    if "Heatmap_Month_Hour" in wb.sheetnames:
        ws = wb["Heatmap_Month_Hour"]
        if ws.max_row > 1 and ws.max_column > 3:
            ws.conditional_formatting.add(
                f"C2:{get_column_letter(ws.max_column)}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

    if include_charts:
        _add_charts(wb)
    if hide_chart_data and "Chart_Data" in wb.sheetnames:
        wb["Chart_Data"].sheet_state = "hidden"
    wb.save(xlsx)


def _add_charts(wb) -> None:
    if "Charts" not in wb.sheetnames or "Chart_Data" not in wb.sheetnames:
        return
    ws = wb["Charts"]
    data = wb["Chart_Data"]
    ws["A1"] = "Prebuilt validation charts"
    ws["A1"].font = Font(bold=True, size=14)

    sections = {}
    for row in range(1, data.max_row + 1):
        value = data.cell(row=row, column=1).value
        if isinstance(value, str) and value.startswith("section:"):
            sections[value.replace("section:", "")] = row

    row = sections.get("chart_monthly_duck")
    if row:
        header = row + 1
        start = row + 2
        end = start + 23
        chart = LineChart()
        chart.title = "Duck curve monthly - weighted mean"
        chart.y_axis.title = "EUR/MWh"
        chart.x_axis.title = "Hour"
        cats = Reference(data, min_col=1, min_row=start, max_row=end)
        values = Reference(data, min_col=2, max_col=data.max_column, min_row=header, max_row=end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 22
        ws.add_chart(chart, "A3")

    row = sections.get("chart_seasonal_duck")
    if row:
        header = row + 1
        start = row + 2
        end = start + 23
        chart = LineChart()
        chart.title = "Duck curve seasonal - weighted mean"
        chart.y_axis.title = "EUR/MWh"
        chart.x_axis.title = "Hour"
        cats = Reference(data, min_col=1, min_row=start, max_row=end)
        values = Reference(data, min_col=2, max_col=5, min_row=header, max_row=end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 22
        ws.add_chart(chart, "A23")

    row = sections.get("chart_width_month")
    if row:
        header = row + 1
        start = row + 2
        end = start + 11
        chart = BarChart()
        chart.title = "Structural width by month"
        chart.y_axis.title = "EUR/MWh"
        chart.x_axis.title = "Month"
        cats = Reference(data, min_col=1, min_row=start, max_row=end)
        values = Reference(data, min_col=2, min_row=header, max_row=end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 15
        ws.add_chart(chart, "Y3")

    row = sections.get("chart_negative_month")
    if row:
        header = row + 1
        start = row + 2
        end = start + 11
        chart = BarChart()
        chart.title = "Negative hours by month"
        chart.y_axis.title = "Hours"
        chart.x_axis.title = "Month"
        cats = Reference(data, min_col=1, min_row=start, max_row=end)
        values = Reference(data, min_col=2, max_col=3, min_row=header, max_row=end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 15
        ws.add_chart(chart, "Y20")


def build_workbook(
    csv_path: Path,
    forwards_path: Path,
    output: Path,
    report: Path | None = None,
    *,
    safe_mode: bool = False,
    include_tables: bool | None = None,
    include_charts: bool | None = None,
    hide_chart_data: bool = True,
) -> Path:
    df, forwards, metrics = _prepare(csv_path, forwards_path)
    annual, residuals, audit_metrics = audit(csv_path, forwards_path)
    seasonal_result = seasonal_audit(csv_path, forwards_path)
    seasonal_checks = seasonal_result["seasonal_checks"]
    quoted_residuals = seasonal_result["quoted_residuals"]
    seasonal_critical = int((seasonal_checks["severity"] == "critical").sum()) if not seasonal_checks.empty else 0
    seasonal_warning = int((seasonal_checks["severity"] == "warning").sum()) if not seasonal_checks.empty else 0
    raw_cols = [
        "timestamp_ch",
        "utc_offset_ch",
        "timestamp_utc",
        "year",
        "quarter",
        "month",
        "season",
        "hour",
        "day_type",
        "calibration_product",
        *PRICE_COLUMNS,
        "negative_weighted",
        "negative_p10",
        "very_low_weighted_lt_5",
    ]
    raw = df[raw_cols].copy()
    raw["timestamp_ch"] = df["ts"].dt.tz_localize(None)
    raw["timestamp_utc"] = df["ts"].dt.tz_convert("UTC").dt.tz_localize(None)
    chart_data = _chart_data(df)
    output.parent.mkdir(parents=True, exist_ok=True)
    use_tables = (not safe_mode) if include_tables is None else bool(include_tables)
    use_charts = (not safe_mode) if include_charts is None else bool(include_charts)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"item": "source_csv", "value": str(csv_path)},
                {"item": "latest_eex_ch_base_date", "value": metrics["latest_eex_ch_base_date"]},
                {"item": "shape_audit_score", "value": audit_metrics["score_10"]},
                {"item": "seasonal_critical_flags", "value": seasonal_critical},
                {"item": "seasonal_warning_flags", "value": seasonal_warning},
                {"item": "production_approval", "value": "NO"},
                {"item": "timestamp_format", "value": "dd.mm.yyyy hh:mm"},
            ]
        ).to_excel(writer, sheet_name="README", index=False)
        raw.to_excel(writer, sheet_name="Raw", index=False)
        _eex_controls(df, forwards).to_excel(writer, sheet_name="EEX_Means", index=False)
        _period_controls(df).to_excel(writer, sheet_name="Period_Means", index=False)
        _duck_month(df).to_excel(writer, sheet_name="Duck_Month", index=False)
        _duck_season(df).to_excel(writer, sheet_name="Duck_Season", index=False)
        _heatmap_month_hour(df).to_excel(writer, sheet_name="Heatmap_Month_Hour", index=False)
        _structural_width(df).to_excel(writer, sheet_name="Structural_Width", index=False)
        _negative_hours(df).to_excel(writer, sheet_name="Negative_Low_Hours", index=False)
        seasonal_checks.to_excel(writer, sheet_name="Seasonal_Coherence", index=False)
        quoted_residuals.to_excel(writer, sheet_name="Quoted_EEX_Products", index=False)
        annual.to_excel(writer, sheet_name="Annual_Audit", index=False)
        residuals.to_excel(writer, sheet_name="Audit_EEX_Residuals", index=False)
        _write_chart_data(writer, chart_data)
        pd.DataFrame(
            {
                "note": [
                    "Safe workbook mode: charts are omitted; use the Duck/Heatmap sheets to create charts in Excel."
                    if safe_mode
                    else "Charts are generated from Chart_Data."
                ]
            }
        ).to_excel(
            writer,
            sheet_name="Charts",
            index=False,
        )
    _style_workbook(
        output,
        include_tables=use_tables,
        include_charts=use_charts,
        hide_chart_data=hide_chart_data,
    )
    if report is not None:
        _write_report(report, workbook=output, csv_path=csv_path, metrics=metrics)
    return output


def _write_chart_data(writer: pd.ExcelWriter, chart_data: dict[str, pd.DataFrame]) -> None:
    sheet = "Chart_Data"
    startrow = 0
    for name, frame in chart_data.items():
        pd.DataFrame([[f"section:{name}"]]).to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=startrow)
        frame.to_excel(writer, sheet_name=sheet, index=False, startrow=startrow + 1)
        startrow += len(frame) + 4


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", default=str(DEFAULT_CSV))
    parser.add_argument("--forwards", default="data/eex_forwards_history.parquet")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    parser.add_argument(
        "--safe-mode",
        action="store_true",
        help="Avoid embedded Excel tables/charts for maximum compatibility.",
    )
    parser.add_argument(
        "--no-tables",
        action="store_true",
        help="Do not create Excel structured tables.",
    )
    parser.add_argument(
        "--no-charts",
        action="store_true",
        help="Do not create embedded charts.",
    )
    parser.add_argument(
        "--show-chart-data",
        action="store_true",
        help="Keep Chart_Data visible for easier Excel chart debugging.",
    )
    args = parser.parse_args(argv)
    out = build_workbook(
        Path(args.csv),
        Path(args.forwards),
        Path(args.output),
        Path(args.report),
        safe_mode=bool(args.safe_mode),
        include_tables=False if args.no_tables else None,
        include_charts=False if args.no_charts else None,
        hide_chart_data=not bool(args.show_chart_data),
    )
    print(f"[validation-workbook] output -> {out}")
    print(f"[validation-workbook] report -> {args.report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
