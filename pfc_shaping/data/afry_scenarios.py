"""Governed local ingestion for restricted AFRY Swiss scenario workbooks.

The workbooks are presentation artifacts, not model-ready tables.  This module
captures and verifies their bytes, rejects worksheet formulas and active
content, disables external-link loading, and publishes typed local Parquet
artifacts with explicit non-authority metadata.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import os
import platform
import re
import shutil
import stat
import sys
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path, PurePosixPath

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

CATALOG_SCHEMA = "fmv-afry-local-catalog-v3"
REGISTRATION_SCHEMA = "fmv-afry-source-registration-v2"
RELEASE_ID = "AFRY_CH_2026_Q2_V100"
CATALOG_DECISION_STATE = "NO_GO_MODEL_AND_PRODUCTION"
CATALOG_INTEGRITY_PASS = "PASS_LOCAL_INTEGRITY_ONLY"
CATALOG_INTEGRITY_NO_GO = "NO_GO_LOCAL_INTEGRITY"
SCENARIOS = ("High", "Central", "Low", "Delayed Transition")
WEATHER_YEARS = (2012, 2014, 2015, 2017, 2018)
PUBLISHED_NEGATIVE_ZERO_TOLERANCE_EUR_MWH = 5e-6
NEAR_DELIVERY_YEARS = tuple(range(2027, 2035))
FAR_DELIVERY_YEARS = tuple(range(2035, 2046)) + (2050, 2055, 2060)
MAX_XLSX_MEMBERS = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 3_000_000_000
MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES = 512_000_000
MAX_XLSX_COMPRESSION_RATIO = 100.0
MAX_SUPPORTING_PDF_BYTES = 100_000_000
INTERPOLATED_FONT_THEME = 0
INTERPOLATED_FONT_TINT = -0.499984740745262
ANNUAL_ENERGY_BALANCE_TOLERANCE_TWH = 0.01
NET_FLOW_ROUNDING_TOLERANCE_TWH = 2.1e-5
HOUR_SLOT_ANNUAL_MEAN_TOLERANCE_EUR_MWH = 1e-10
HOUR_SLOT_ZERO_MEAN_TOLERANCE_EUR_MWH = 1e-12
SWISS_2035_NON_HYDRO_RENEWABLE_TARGET_TWH = 35.0
AFRY_NON_HYDRO_RENEWABLE_PROXY_TECHNOLOGIES = (
    "Solar PV",
    "Onshore Wind",
    "Other Res",
    "CCS Biomass",
)
HOURLY_HEADER = (
    "Price area",
    "Scenario",
    "Year",
    "Weather year",
    "Hour",
    "Wholesale price",
)

CATALOG_INTEGRITY_GATE_NAMES = (
    "source_hashes_exact",
    "source_capture_stable",
    "workbook_containers_safe",
    "semantic_contract_exact",
    "hourly_schema_exact",
    "hourly_group_coverage_exact",
    "representative_hour_continuity",
    "annual_semantic_reconciliation",
    "structural_contract_complete",
    "annual_hour_slot_shape_zero_mean",
    "hourly_annual_rounding_reconciliation",
    "negative_hour_count_reconciliation",
)


class AfryDataContractError(ValueError):
    """Raised when a source or extracted table violates the AFRY contract."""


@dataclass(frozen=True)
class AfrySource:
    role: str
    path: Path
    sha256: str
    publisher_release_date: str
    declared_money_basis: str = ""


@dataclass(frozen=True)
class AfryCapturedSource:
    role: str
    payload: bytes
    sha256: str
    source_name: str
    publisher_release_date: str
    declared_money_basis: str = ""


@dataclass(frozen=True)
class AfryRegistration:
    release_id: str
    first_observed_at_utc: str
    cadence: str
    expected_arrival_month: int
    schedule_timezone: str
    sources: tuple[AfrySource, ...]
    supporting_documents: tuple[AfrySupportingDocument, ...]
    semantic_contract_path: Path
    semantic_contract_sha256: str
    hourly_contract: AfryHourlyContract
    annual_contract: AfryAnnualContract


@dataclass(frozen=True)
class AfryHourlySheetContract:
    title: str
    scenario_vendor: str
    delivery_years: tuple[int, ...]


@dataclass(frozen=True)
class AfryHourlyContract:
    price_area: str
    weather_years: tuple[int, ...]
    representative_hours_per_group: int
    negative_zero_tolerance_eur_mwh: float
    sheets: tuple[AfryHourlySheetContract, ...]

    @property
    def expected_total_rows(self) -> int:
        return sum(len(sheet.delivery_years) for sheet in self.sheets) * len(
            self.weather_years
        ) * self.representative_hours_per_group


@dataclass(frozen=True)
class AfryAnnualContract:
    release_year: int
    first_delivery_year_offset: int
    delivery_year_count: int
    delivery_years: tuple[int, ...]
    structural_required_non_null_fields: tuple[str, ...]
    structural_declared_nullable_fields: tuple[str, ...]


@dataclass(frozen=True)
class AfrySupportingDocument:
    role: str
    path: Path
    sha256: str
    publisher_release_date: str


HOURLY_SCHEMA = pa.schema(
    [
        pa.field("release_id", pa.string(), nullable=False),
        pa.field("price_area", pa.string(), nullable=False),
        pa.field("scenario_vendor", pa.string(), nullable=False),
        pa.field("delivery_year", pa.int16(), nullable=False),
        pa.field("weather_year", pa.int16(), nullable=False),
        pa.field("representative_hour", pa.int16(), nullable=False),
        pa.field("price_eur_mwh_real_release_basis", pa.float64(), nullable=False),
        pa.field("source_sheet", pa.string(), nullable=False),
        pa.field("source_row", pa.int32(), nullable=False),
    ],
    metadata={
        b"time_semantics": b"representative_hour_1_to_8760_not_utc_timestamp",
        b"level_authority": b"false",
        b"model_input_authorized": b"false",
    },
)

HOUR_SLOT_SHAPE_SCHEMA = pa.schema(
    [
        pa.field("release_id", pa.string(), nullable=False),
        pa.field("scenario_vendor", pa.string(), nullable=False),
        pa.field("delivery_year", pa.int16(), nullable=False),
        pa.field("weather_year", pa.int16(), nullable=False),
        pa.field("representative_hour_slot", pa.int8(), nullable=False),
        pa.field("annual_mean_eur_mwh_real_release_basis", pa.float64(), nullable=False),
        pa.field("hour_slot_mean_eur_mwh_real_release_basis", pa.float64(), nullable=False),
        pa.field("annual_zero_mean_additive_shape_eur_mwh", pa.float64(), nullable=False),
        pa.field("multiplicative_factor", pa.float64()),
        pa.field("model_input_authorized", pa.bool_(), nullable=False),
        pa.field("monthly_level_authority", pa.bool_(), nullable=False),
        pa.field("production_authorized", pa.bool_(), nullable=False),
    ],
    metadata={
        b"time_semantics": b"representative_hour_slot_0_to_23_not_local_or_utc_timestamp",
        b"normalization": b"zero_mean_across_24_slots_within_scenario_delivery_weather_group",
        b"monthly_zero_mean_guaranteed": b"false",
        b"model_input_authorized": b"false",
        b"monthly_level_authority": b"false",
        b"production_authorized": b"false",
    },
)

ANNUAL_FACT_SCHEMA = pa.schema(
    [
        pa.field("release_id", pa.string(), nullable=False),
        pa.field("sheet", pa.string(), nullable=False),
        pa.field("table_header_row", pa.int16(), nullable=False),
        pa.field("table_title", pa.string(), nullable=False),
        pa.field("unit_raw", pa.string(), nullable=False),
        pa.field("scenario_vendor", pa.string()),
        pa.field("dimension_1_raw", pa.string()),
        pa.field("dimension_2_raw", pa.string()),
        pa.field("dimension_3_raw", pa.string()),
        pa.field("dimension_1", pa.string()),
        pa.field("dimension_2", pa.string()),
        pa.field("dimension_3", pa.string()),
        pa.field("delivery_year", pa.int16(), nullable=False),
        pa.field("value", pa.float64(), nullable=False),
        pa.field("value_status", pa.string(), nullable=False),
        pa.field("source_cell", pa.string(), nullable=False),
    ],
    metadata={
        b"value_semantics": b"vendor_value_with_unit_raw_no_implicit_unit_conversion",
        b"font_semantics": b"historical_or_vendor_black_modelled_or_vendor_grey_interpolated",
        b"level_authority": b"false",
        b"model_input_authorized": b"false",
    },
)

STRUCTURAL_SCHEMA = pa.schema(
    [
        pa.field("publication_date", pa.timestamp("us", tz="UTC"), nullable=False),
        pa.field("measurement_date", pa.timestamp("us", tz="UTC")),
        pa.field("ingested_at_utc", pa.timestamp("us", tz="UTC"), nullable=False),
        pa.field("source", pa.string(), nullable=False),
        pa.field("scenario", pa.string(), nullable=False),
        pa.field("scenario_edition", pa.string(), nullable=False),
        pa.field("country", pa.string(), nullable=False),
        pa.field("delivery_year", pa.int16(), nullable=False),
        pa.field("delivery_month", pa.int8()),
        pa.field("track", pa.string(), nullable=False),
        pa.field("quality_flag", pa.string(), nullable=False),
        pa.field("source_value_statuses", pa.string(), nullable=False),
        pa.field("field_value_statuses_json", pa.string(), nullable=False),
        pa.field("contains_interpolated_values", pa.bool_(), nullable=False),
        pa.field("scenario_weight", pa.float64()),
        pa.field("demand_twh", pa.float64()),
        pa.field("peak_load_gw", pa.float64()),
        pa.field("pv_gw", pa.float64()),
        pa.field("pv_twh", pa.float64()),
        pa.field("wind_gw", pa.float64()),
        pa.field("wind_twh", pa.float64()),
        pa.field("battery_power_gw", pa.float64()),
        pa.field("battery_energy_gwh", pa.float64()),
        pa.field("electrolysis_twh", pa.float64()),
        pa.field("hydro_twh", pa.float64()),
        pa.field("hydro_capacity_gw", pa.float64()),
        pa.field("hydro_reservoir_twh", pa.float64()),
        pa.field("nuclear_gw", pa.float64()),
        pa.field("gas_gw", pa.float64()),
        pa.field("net_import_twh", pa.float64()),
        pa.field("gas_eur_mwh", pa.float64()),
        pa.field("coal_eur_mwh", pa.float64()),
        pa.field("co2_eur_t", pa.float64()),
        pa.field("model_input_authorized", pa.bool_(), nullable=False),
        pa.field("monthly_level_authority", pa.bool_(), nullable=False),
        pa.field("production_authorized", pa.bool_(), nullable=False),
    ],
    metadata={
        b"scenario_semantics": b"vendor_labels_unmapped_no_probabilities",
        b"usage": b"restricted_benchmark_only",
        b"monthly_level_authority": b"false",
        b"model_input_authorized": b"false",
        b"production_authorized": b"false",
    },
)

STRUCTURAL_REQUIRED_NON_NULL_FIELDS = (
    "demand_twh",
    "peak_load_gw",
    "pv_gw",
    "pv_twh",
    "wind_gw",
    "wind_twh",
    "battery_power_gw",
    "electrolysis_twh",
    "hydro_twh",
    "hydro_capacity_gw",
    "nuclear_gw",
    "gas_gw",
    "net_import_twh",
    "gas_eur_mwh",
    "coal_eur_mwh",
    "co2_eur_t",
)
STRUCTURAL_DECLARED_NULLABLE_FIELDS = (
    "measurement_date",
    "delivery_month",
    "scenario_weight",
    "battery_energy_gwh",
    "hydro_reservoir_twh",
)
CATALOG_ARTIFACT_NAMES = frozenset(
    {
        "hourly_prices.parquet",
        "annual_facts.parquet",
        "structural_scenarios_benchmark.parquet",
        "hour_slot_shape_benchmark.parquet",
        "hourly_profile.json",
        "annual_semantic_profile.json",
        "structural_contract_audit.json",
        "swiss_2035_policy_gap_profile.json",
        "hour_slot_shape_audit.json",
    }
)
CATALOG_PARQUET_SCHEMAS = {
    "hourly_prices.parquet": HOURLY_SCHEMA,
    "annual_facts.parquet": ANNUAL_FACT_SCHEMA,
    "structural_scenarios_benchmark.parquet": STRUCTURAL_SCHEMA,
    "hour_slot_shape_benchmark.parquet": HOUR_SLOT_SHAPE_SCHEMA,
}


def canonical_json_bytes(value: object) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n"
    ).encode("utf-8")


def catalog_content_id(catalog_body: dict[str, object]) -> str:
    """Hash every authoritative catalog field except the identifier itself."""

    if "bundle_id" in catalog_body:
        raise AfryDataContractError("catalog body must not contain bundle_id before hashing")
    return hashlib.sha256(canonical_json_bytes(catalog_body)).hexdigest()


def _verify_catalog_policy(catalog: dict[str, object]) -> None:
    """Reject self-consistent catalogs that attempt to grant AFRY authority."""

    if catalog.get("schema_version") != CATALOG_SCHEMA:
        raise AfryDataContractError("unsupported AFRY catalog schema")
    if catalog.get("decision_state") != CATALOG_DECISION_STATE:
        raise AfryDataContractError(
            f"AFRY catalog decision_state must remain {CATALOG_DECISION_STATE}"
        )
    admission = catalog.get("admission")
    if not isinstance(admission, dict):
        raise AfryDataContractError("AFRY catalog admission must be an object")
    required_admission = {
        "license_classification": "restricted_client_engagement",
        "raw_or_derived_values_allowed_in_git": False,
        "rag_indexing_allowed": False,
        "monthly_level_authority": False,
        "shaping_benchmark_only": True,
        "scenario_mapping_approved": False,
        "model_input_authorized": False,
        "production_authorized": False,
    }
    for field, expected in required_admission.items():
        if admission.get(field) != expected:
            raise AfryDataContractError(
                f"AFRY catalog admission invariant {field} must be {expected!r}"
            )
    gates = catalog.get("gates")
    if not isinstance(gates, dict):
        raise AfryDataContractError("AFRY catalog gates must be an object")
    for field in ("model_input_authorized", "production_authorized"):
        if gates.get(field) != "NO_GO":
            raise AfryDataContractError(f"AFRY catalog gate {field} must remain NO_GO")
    if not str(gates.get("swiss_2035_policy_target_reconciliation", "")).startswith(
        "NO_GO"
    ):
        raise AfryDataContractError(
            "AFRY Swiss policy reconciliation gate must remain NO_GO"
        )
    expected_integrity = (
        CATALOG_INTEGRITY_PASS
        if all(gates.get(name) == "PASS" for name in CATALOG_INTEGRITY_GATE_NAMES)
        else CATALOG_INTEGRITY_NO_GO
    )
    if catalog.get("integrity_verdict") != expected_integrity:
        raise AfryDataContractError(
            "AFRY catalog integrity_verdict does not match its local integrity gates"
        )
    if "qa_verdict" in catalog:
        raise AfryDataContractError(
            "AFRY catalog v2 forbids ambiguous qa_verdict; use integrity_verdict"
        )
    registration = catalog.get("registration")
    if not isinstance(registration, dict):
        raise AfryDataContractError("AFRY catalog registration receipt must be an object")
    registration_path = registration.get("path")
    if (
        not isinstance(registration_path, str)
        or Path(registration_path).is_absolute()
        or ".." in PurePosixPath(registration_path.replace("\\", "/")).parts
        or "\\" in registration_path
        or registration.get("schema_version") != REGISTRATION_SCHEMA
        or not re.fullmatch(r"[0-9a-f]{64}", str(registration.get("sha256", "")))
        or not isinstance(registration.get("bytes"), int)
        or registration["bytes"] <= 0
        or registration.get("stable_during_materialization") is not True
    ):
        raise AfryDataContractError("AFRY catalog registration receipt is invalid")
    annual_contract = catalog.get("annual_contract")
    release_match = re.fullmatch(
        r"AFRY_CH_(\d{4})_Q[1-4]_V\d+", str(catalog.get("release_id", ""))
    )
    if not isinstance(annual_contract, dict) or release_match is None:
        raise AfryDataContractError("AFRY catalog annual contract is missing")
    release_year = int(release_match.group(1))
    first_offset = annual_contract.get("first_delivery_year_offset")
    year_count = annual_contract.get("delivery_year_count")
    delivery_years = annual_contract.get("delivery_years")
    if (
        annual_contract.get("release_year") != release_year
        or first_offset != 1
        or type(year_count) is not int
        or not 1 <= year_count <= 100
        or not isinstance(delivery_years, list)
        or delivery_years
        != list(range(release_year + first_offset, release_year + first_offset + year_count))
        or annual_contract.get("scenarios") != list(SCENARIOS)
    ):
        raise AfryDataContractError("AFRY catalog annual contract is invalid")
    producer = catalog.get("producer")
    if not isinstance(producer, dict):
        raise AfryDataContractError("AFRY catalog producer provenance must be an object")
    files = producer.get("files")
    if (
        not isinstance(files, list)
        or len(files) != 2
        or {item.get("role") for item in files if isinstance(item, dict)}
        != {"ingestion_module", "materializer_entrypoint"}
    ):
        raise AfryDataContractError("AFRY catalog producer files are incomplete")
    for item in files:
        path_value = item.get("path") if isinstance(item, dict) else None
        if (
            not isinstance(item, dict)
            or not isinstance(path_value, str)
            or Path(path_value).is_absolute()
            or ".." in PurePosixPath(path_value.replace("\\", "/")).parts
            or "\\" in path_value
            or not re.fullmatch(r"[0-9a-f]{64}", str(item.get("sha256", "")))
        ):
            raise AfryDataContractError("AFRY catalog producer file provenance is invalid")
    runtime = producer.get("runtime")
    if not isinstance(runtime, dict) or any(
        not isinstance(runtime.get(name), str) or not runtime[name]
        for name in (
            "python",
            "implementation",
            "cache_tag",
            "platform_system",
            "platform_machine",
            "executable_path",
            "executable_sha256",
            "prefix_path",
            "pandas",
            "pyarrow",
            "openpyxl",
        )
    ):
        raise AfryDataContractError("AFRY catalog producer runtime provenance is incomplete")
    executable_path = str(runtime["executable_path"])
    prefix_path = str(runtime["prefix_path"])
    if (
        Path(executable_path).is_absolute()
        or ".." in PurePosixPath(executable_path.replace("\\", "/")).parts
        or "\\" in executable_path
        or Path(prefix_path).is_absolute()
        or ".." in PurePosixPath(prefix_path.replace("\\", "/")).parts
        or "\\" in prefix_path
        or PurePosixPath(executable_path).parent != PurePosixPath(prefix_path)
        or not re.fullmatch(r"[0-9a-f]{64}", str(runtime["executable_sha256"]))
    ):
        raise AfryDataContractError("AFRY catalog runtime executable provenance is invalid")
    closure = runtime.get("dependency_closure")
    if (
        not isinstance(closure, dict)
        or closure.get("schema_version") != "fmv-afry-runtime-tree-receipt-v1"
        or re.fullmatch(r"[0-9a-f]{64}", str(closure.get("tree_sha256", ""))) is None
        or type(closure.get("file_count")) is not int
        or closure["file_count"] <= 0
        or type(closure.get("total_bytes")) is not int
        or closure["total_bytes"] <= 0
        or closure.get("plain_file_tree") is not True
        or closure.get("stable_during_materialization") is not True
    ):
        raise AfryDataContractError("AFRY catalog runtime dependency closure is invalid")
    if producer.get("source_stable_during_materialization") is not True:
        raise AfryDataContractError("AFRY producer stability evidence must be true")


def verify_catalog_bundle(bundle: Path) -> dict[str, object]:
    """Verify the complete member set and content identity of a published bundle."""

    if not bundle.is_dir() or bundle.is_symlink():
        raise AfryDataContractError(f"AFRY catalog bundle is not a plain directory: {bundle}")
    catalog_path = bundle / "catalog.json"
    checksum_path = bundle / "catalog.sha256"
    if not catalog_path.is_file() or catalog_path.is_symlink():
        raise AfryDataContractError(f"AFRY catalog manifest is missing or linked: {catalog_path}")
    try:
        catalog_bytes = _stable_regular_file_bytes(catalog_path)
        catalog = json.loads(catalog_bytes)
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise AfryDataContractError(f"AFRY catalog manifest is unreadable: {catalog_path}") from exc
    if not isinstance(catalog, dict):
        raise AfryDataContractError("AFRY catalog manifest must be a JSON object")
    bundle_id = catalog.get("bundle_id")
    if not isinstance(bundle_id, str):
        raise AfryDataContractError("AFRY catalog bundle_id is missing")
    body = dict(catalog)
    body.pop("bundle_id")
    if catalog_content_id(body) != bundle_id or bundle.name != bundle_id:
        raise AfryDataContractError("AFRY catalog content identity mismatch")
    _verify_catalog_policy(catalog)
    artifacts = catalog.get("artifacts")
    if not isinstance(artifacts, list):
        raise AfryDataContractError("AFRY catalog artifacts must be a list")
    artifact_by_name: dict[str, dict[str, object]] = {}
    for artifact in artifacts:
        if not isinstance(artifact, dict):
            raise AfryDataContractError("AFRY catalog artifact entry must be an object")
        name = artifact.get("name")
        if (
            not isinstance(name, str)
            or not name
            or Path(name).name != name
            or "/" in name
            or "\\" in name
            or name in artifact_by_name
        ):
            raise AfryDataContractError(f"unsafe or duplicate AFRY artifact name: {name!r}")
        artifact_by_name[name] = artifact
    if set(artifact_by_name) != CATALOG_ARTIFACT_NAMES:
        raise AfryDataContractError(
            "AFRY catalog artifact policy closure mismatch: "
            f"expected={sorted(CATALOG_ARTIFACT_NAMES)} "
            f"actual={sorted(artifact_by_name)}"
        )
    expected_members = set(artifact_by_name) | {"catalog.json", "catalog.sha256"}
    entries = list(bundle.iterdir())
    actual_members = {entry.name for entry in entries}
    if actual_members != expected_members or len(entries) != len(expected_members):
        raise AfryDataContractError(
            "AFRY catalog member closure mismatch: "
            f"expected={sorted(expected_members)} actual={sorted(actual_members)}"
        )
    if any(not entry.is_file() or entry.is_symlink() for entry in entries):
        raise AfryDataContractError("AFRY catalog members must be plain files")
    expected_checksum = f"{hashlib.sha256(catalog_bytes).hexdigest()}\n".encode("ascii")
    if _stable_regular_file_bytes(checksum_path) != expected_checksum:
        raise AfryDataContractError("AFRY catalog manifest checksum mismatch")
    artifact_payloads: dict[str, bytes] = {}
    for name, artifact in artifact_by_name.items():
        path = bundle / name
        payload = _stable_regular_file_bytes(path)
        if len(payload) != artifact.get("bytes"):
            raise AfryDataContractError(f"AFRY catalog artifact size mismatch: {name}")
        if hashlib.sha256(payload).hexdigest() != artifact.get("sha256"):
            raise AfryDataContractError(f"AFRY catalog artifact sha256 mismatch: {name}")
        artifact_payloads[name] = payload
    for name, expected_schema in CATALOG_PARQUET_SCHEMAS.items():
        parquet = pq.ParquetFile(pa.BufferReader(artifact_payloads[name]))
        if not parquet.schema_arrow.equals(expected_schema, check_metadata=True):
            raise AfryDataContractError(f"AFRY catalog Parquet schema mismatch: {name}")
        declared_rows = artifact_by_name[name].get("rows")
        if (
            not isinstance(declared_rows, int)
            or declared_rows <= 0
            or parquet.metadata.num_rows != declared_rows
        ):
            raise AfryDataContractError(f"AFRY catalog Parquet row-count mismatch: {name}")
        authority_columns = [
            field
            for field in (
                "model_input_authorized",
                "monthly_level_authority",
                "production_authorized",
            )
            if field in expected_schema.names
        ]
        if authority_columns:
            table = pq.read_table(
                pa.BufferReader(artifact_payloads[name]), columns=authority_columns
            )
            for field in authority_columns:
                if set(table[field].to_pylist()) != {False}:
                    raise AfryDataContractError(
                        f"AFRY catalog Parquet authority must remain false: {name}/{field}"
                    )
    for name in (
        "structural_contract_audit.json",
        "swiss_2035_policy_gap_profile.json",
        "hour_slot_shape_audit.json",
    ):
        payload = json.loads(artifact_payloads[name])
        for field in (
            "model_input_authorized",
            "monthly_level_authority",
            "production_authorized",
        ):
            if payload.get(field) is not False:
                raise AfryDataContractError(
                    f"AFRY catalog JSON authority must remain false: {name}/{field}"
                )
    annual_contract_payload = catalog["annual_contract"]
    annual_contract = AfryAnnualContract(
        release_year=int(annual_contract_payload["release_year"]),
        first_delivery_year_offset=int(
            annual_contract_payload["first_delivery_year_offset"]
        ),
        delivery_year_count=int(annual_contract_payload["delivery_year_count"]),
        delivery_years=tuple(annual_contract_payload["delivery_years"]),
        structural_required_non_null_fields=STRUCTURAL_REQUIRED_NON_NULL_FIELDS,
        structural_declared_nullable_fields=STRUCTURAL_DECLARED_NULLABLE_FIELDS,
    )
    structural = pq.read_table(
        pa.BufferReader(artifact_payloads["structural_scenarios_benchmark.parquet"])
    ).to_pandas()
    recomputed_structural_audit = validate_structural_contract(
        structural, annual_contract
    )
    declared_structural_audit = json.loads(
        artifact_payloads["structural_contract_audit.json"]
    )
    if (
        not recomputed_structural_audit["all_checks_pass"]
        or declared_structural_audit != recomputed_structural_audit
    ):
        raise AfryDataContractError(
            "AFRY catalog structural audit does not match Parquet bytes"
        )
    annual_facts = pq.read_table(
        pa.BufferReader(artifact_payloads["annual_facts.parquet"])
    ).to_pandas()
    recomputed_annual_semantics = validate_annual_semantics(
        annual_facts,
        delivery_years=annual_contract.delivery_years,
        release_year=annual_contract.release_year,
    )
    declared_annual_semantics = json.loads(
        artifact_payloads["annual_semantic_profile.json"]
    )
    if (
        not recomputed_annual_semantics["all_checks_pass"]
        or declared_annual_semantics != recomputed_annual_semantics
    ):
        raise AfryDataContractError(
            "AFRY catalog annual semantic audit does not match Parquet bytes"
        )
    return catalog


def read_verified_artifact_bytes(
    bundle: Path, artifact_name: str
) -> tuple[dict[str, object], bytes]:
    """Verify a bundle, then single-read and rebind one artifact for consumption."""

    catalog = verify_catalog_bundle(bundle)
    artifacts = catalog["artifacts"]
    artifact = next(
        (
            item
            for item in artifacts
            if isinstance(item, dict) and item.get("name") == artifact_name
        ),
        None,
    )
    if artifact is None:
        raise AfryDataContractError(f"unknown AFRY catalog artifact: {artifact_name!r}")
    payload = _stable_regular_file_bytes(bundle / artifact_name)
    if (
        len(payload) != artifact.get("bytes")
        or hashlib.sha256(payload).hexdigest() != artifact.get("sha256")
    ):
        raise AfryDataContractError(
            f"AFRY artifact changed between bundle verification and capture: {artifact_name}"
        )
    return catalog, payload


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _is_link_or_reparse(path: Path, path_stat: os.stat_result) -> bool:
    attributes = int(getattr(path_stat, "st_file_attributes", 0))
    reparse_flag = int(getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400))
    return path.is_symlink() or bool(attributes & reparse_flag)


def _stable_regular_file_bytes(path: Path) -> bytes:
    before = path.stat(follow_symlinks=False)
    if _is_link_or_reparse(path, before) or not stat.S_ISREG(before.st_mode):
        raise AfryDataContractError(f"AFRY file is not plain: {path}")
    identity_before = (
        int(before.st_dev),
        int(before.st_ino),
        int(before.st_size),
        int(before.st_mtime_ns),
        int(before.st_nlink),
    )
    payload = path.read_bytes()
    after = path.stat(follow_symlinks=False)
    identity_after = (
        int(after.st_dev),
        int(after.st_ino),
        int(after.st_size),
        int(after.st_mtime_ns),
        int(after.st_nlink),
    )
    if identity_before != identity_after or len(payload) != before.st_size:
        raise AfryDataContractError(f"AFRY plain file changed while read: {path}")
    return payload


def _stable_regular_file_receipt(path: Path) -> tuple[str, int]:
    payload = _stable_regular_file_bytes(path)
    return hashlib.sha256(payload).hexdigest(), len(payload)


def runtime_tree_receipt(prefix: Path) -> dict[str, object]:
    """Bind every byte in the repo-local AFRY intake runtime prefix."""

    prefix = prefix.resolve()
    prefix_stat = prefix.stat(follow_symlinks=False)
    if _is_link_or_reparse(prefix, prefix_stat) or not stat.S_ISDIR(prefix_stat.st_mode):
        raise AfryDataContractError(f"AFRY runtime prefix is not a plain directory: {prefix}")
    files: list[tuple[str, str, int]] = []
    for path in prefix.rglob("*"):
        path_stat = path.stat(follow_symlinks=False)
        if _is_link_or_reparse(path, path_stat):
            raise AfryDataContractError(f"AFRY runtime closure contains a link: {path}")
        if stat.S_ISDIR(path_stat.st_mode):
            continue
        sha256, byte_count = _stable_regular_file_receipt(path)
        files.append((path.relative_to(prefix).as_posix(), sha256, byte_count))
    if not files:
        raise AfryDataContractError("AFRY runtime closure must contain files")
    digest = hashlib.sha256()
    for relative, sha256, byte_count in sorted(files):
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(sha256.encode("ascii"))
        digest.update(b"\0")
        digest.update(str(byte_count).encode("ascii"))
        digest.update(b"\0")
    return {
        "schema_version": "fmv-afry-runtime-tree-receipt-v1",
        "tree_sha256": digest.hexdigest(),
        "file_count": len(files),
        "total_bytes": sum(item[2] for item in files),
        "plain_file_tree": True,
        "stable_during_materialization": True,
    }


def capture_verified_source(
    source: AfrySource,
) -> tuple[AfryCapturedSource, dict[str, object]]:
    """Capture immutable source bytes once before any workbook parsing."""

    if source.role not in {"annual_scenarios", "hourly_prices"}:
        raise AfryDataContractError(f"unsupported AFRY source role: {source.role!r}")
    if not source.path.is_file():
        raise FileNotFoundError(source.path)
    digest = hashlib.sha256()
    byte_count = 0
    chunks: list[bytes] = []
    with source.path.open("rb") as reader:
        for chunk in iter(lambda: reader.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
            chunks.append(chunk)
            byte_count += len(chunk)
    actual = digest.hexdigest()
    if actual != source.sha256:
        raise AfryDataContractError(
            f"AFRY {source.role} capture sha256 mismatch: "
            f"expected={source.sha256} actual={actual}"
        )
    captured = AfryCapturedSource(
        role=source.role,
        payload=b"".join(chunks),
        sha256=source.sha256,
        source_name=source.path.name,
        publisher_release_date=source.publisher_release_date,
        declared_money_basis=source.declared_money_basis,
    )
    receipt = _verify_source_payload(source, captured.payload)
    receipt.update(
        path=source.path.name,
        captured_bytes=byte_count,
        capture_sha256=actual,
        capture_location="immutable_process_memory",
    )
    return captured, receipt


def _parse_utc(value: str) -> str:
    parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        raise AfryDataContractError(f"timestamp must be timezone-aware: {value!r}")
    return parsed.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def _parse_date(value: object, *, field: str) -> date:
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise AfryDataContractError(f"{field} must be an ISO date") from exc


def load_registration(
    path: Path,
    *,
    workspace: Path,
    raw_bytes: bytes | None = None,
) -> AfryRegistration:
    if not path.is_file():
        raise FileNotFoundError(path)
    payload = json.loads((raw_bytes if raw_bytes is not None else path.read_bytes()).decode("utf-8"))
    if payload.get("schema_version") != REGISTRATION_SCHEMA:
        raise AfryDataContractError("unsupported AFRY source registration schema")
    release_id = str(payload.get("release_id", ""))
    release_match = re.fullmatch(r"AFRY_CH_(\d{4})_Q[1-4]_V\d+", release_id)
    if release_match is None:
        raise AfryDataContractError("invalid AFRY release_id")
    release_year = int(release_match.group(1))
    observed_at = _parse_utc(str(payload["first_observed_at_utc"]))
    observed_date = datetime.fromisoformat(observed_at.replace("Z", "+00:00")).date()
    if payload.get("license_classification") != "restricted_client_engagement":
        raise AfryDataContractError("AFRY license_classification must remain restricted")
    for field in (
        "raw_or_derived_values_allowed_in_git",
        "rag_indexing_allowed",
        "monthly_level_authority",
        "scenario_mapping_approved",
        "model_input_authorized",
        "production_authorized",
    ):
        if payload.get(field) is not False:
            raise AfryDataContractError(f"AFRY registration invariant {field} must be false")
    schedule = payload.get("release_schedule")
    if not isinstance(schedule, dict):
        raise AfryDataContractError("AFRY registration missing release_schedule")
    cadence = str(schedule.get("cadence"))
    expected_arrival_month = schedule.get("expected_arrival_month")
    schedule_timezone = str(schedule.get("timezone"))
    if cadence != "annual":
        raise AfryDataContractError("AFRY release cadence must be annual")
    if expected_arrival_month != 7:
        raise AfryDataContractError("AFRY expected arrival month must be July (7)")
    if schedule_timezone != "Europe/Zurich":
        raise AfryDataContractError("AFRY release schedule timezone must be Europe/Zurich")
    hourly_payload = payload.get("content_contract", {}).get("hourly_prices")
    if not isinstance(hourly_payload, dict):
        raise AfryDataContractError("AFRY registration missing hourly content contract")
    price_area = str(hourly_payload.get("price_area", ""))
    if price_area != "SWI":
        raise AfryDataContractError("AFRY hourly price_area must be SWI")
    weather_years = tuple(hourly_payload.get("weather_years", ()))
    if (
        not weather_years
        or any(type(year) is not int or not 1900 <= year <= 2100 for year in weather_years)
        or len(set(weather_years)) != len(weather_years)
    ):
        raise AfryDataContractError("AFRY hourly weather_years must be unique integers")
    representative_hours = hourly_payload.get("representative_hours_per_group")
    if representative_hours != 8_760:
        raise AfryDataContractError(
            "AFRY hourly representative_hours_per_group must remain 8760"
        )
    negative_zero_tolerance = hourly_payload.get("negative_zero_tolerance_eur_mwh")
    if negative_zero_tolerance != PUBLISHED_NEGATIVE_ZERO_TOLERANCE_EUR_MWH:
        raise AfryDataContractError(
            "AFRY hourly negative_zero_tolerance_eur_mwh must be explicitly versioned"
        )
    sheet_contracts: list[AfryHourlySheetContract] = []
    for item in hourly_payload.get("sheets", []):
        title = str(item.get("title", "")).strip()
        scenario = str(item.get("scenario_vendor", "")).strip()
        years = tuple(item.get("delivery_years", ()))
        if not title or scenario not in SCENARIOS:
            raise AfryDataContractError("invalid AFRY hourly sheet title or vendor scenario")
        if (
            not years
            or any(type(year) is not int or not 2020 <= year <= 2100 for year in years)
            or len(set(years)) != len(years)
        ):
            raise AfryDataContractError(
                f"AFRY hourly delivery_years must be unique integers for {title!r}"
            )
        sheet_contracts.append(
            AfryHourlySheetContract(
                title=title,
                scenario_vendor=scenario,
                delivery_years=years,
            )
        )
    if not sheet_contracts or len({sheet.title for sheet in sheet_contracts}) != len(
        sheet_contracts
    ):
        raise AfryDataContractError("AFRY hourly sheet titles must be present and unique")
    years_by_scenario: dict[str, list[int]] = {scenario: [] for scenario in SCENARIOS}
    for sheet in sheet_contracts:
        years_by_scenario[sheet.scenario_vendor].extend(sheet.delivery_years)
    reference_years: tuple[int, ...] | None = None
    for scenario, years in years_by_scenario.items():
        if len(set(years)) != len(years):
            raise AfryDataContractError(
                f"AFRY hourly delivery years overlap across sheets for {scenario}"
            )
        ordered = tuple(sorted(years))
        if not ordered:
            raise AfryDataContractError(f"AFRY hourly scenario has no delivery years: {scenario}")
        if reference_years is None:
            reference_years = ordered
        elif ordered != reference_years:
            raise AfryDataContractError(
                "AFRY hourly scenarios must have identical delivery-year coverage"
            )
    hourly_contract = AfryHourlyContract(
        price_area=price_area,
        weather_years=weather_years,
        representative_hours_per_group=representative_hours,
        negative_zero_tolerance_eur_mwh=float(negative_zero_tolerance),
        sheets=tuple(sheet_contracts),
    )
    annual_payload = payload.get("content_contract", {}).get("annual_scenarios")
    if not isinstance(annual_payload, dict):
        raise AfryDataContractError("AFRY registration missing annual content contract")
    first_delivery_year_offset = annual_payload.get("first_delivery_year_offset")
    delivery_year_count = annual_payload.get("delivery_year_count")
    if first_delivery_year_offset != 1:
        raise AfryDataContractError(
            "AFRY annual first_delivery_year_offset must be exactly one year"
        )
    if type(delivery_year_count) is not int or not 1 <= delivery_year_count <= 100:
        raise AfryDataContractError(
            "AFRY annual delivery_year_count must be an integer in [1, 100]"
        )
    annual_years = tuple(annual_payload.get("delivery_years", ()))
    expected_annual_years = tuple(
        range(
            release_year + first_delivery_year_offset,
            release_year + first_delivery_year_offset + delivery_year_count,
        )
    )
    if annual_years != expected_annual_years:
        raise AfryDataContractError(
            "AFRY annual delivery_years must be the exact contiguous release-anchored horizon"
        )
    if reference_years is None or not set(reference_years).issubset(annual_years):
        raise AfryDataContractError(
            "AFRY hourly delivery years must be a subset of the annual horizon"
        )
    if reference_years[0] != annual_years[0] or reference_years[-1] != annual_years[-1]:
        raise AfryDataContractError(
            "AFRY hourly and annual delivery horizons must share first and last years"
        )
    required_fields = tuple(annual_payload.get("structural_required_non_null_fields", ()))
    nullable_fields = tuple(annual_payload.get("structural_declared_nullable_fields", ()))
    if required_fields != STRUCTURAL_REQUIRED_NON_NULL_FIELDS:
        raise AfryDataContractError(
            "AFRY annual structural_required_non_null_fields contract mismatch"
        )
    if nullable_fields != STRUCTURAL_DECLARED_NULLABLE_FIELDS:
        raise AfryDataContractError(
            "AFRY annual structural_declared_nullable_fields contract mismatch"
        )
    annual_contract = AfryAnnualContract(
        release_year=release_year,
        first_delivery_year_offset=first_delivery_year_offset,
        delivery_year_count=delivery_year_count,
        delivery_years=annual_years,
        structural_required_non_null_fields=required_fields,
        structural_declared_nullable_fields=nullable_fields,
    )
    sources: list[AfrySource] = []
    for item in payload.get("sources", []):
        relative = Path(str(item["path"]))
        if relative.is_absolute() or relative.suffix.lower() != ".xlsx":
            raise AfryDataContractError("AFRY source paths must be relative .xlsx paths")
        source_path = (workspace / relative).resolve()
        try:
            source_path.relative_to(workspace.resolve())
        except ValueError as exc:
            raise AfryDataContractError(
                "AFRY source must be inside the canonical workspace"
            ) from exc
        expected_hash = str(item["sha256"]).lower()
        if re.fullmatch(r"[0-9a-f]{64}", expected_hash) is None:
            raise AfryDataContractError("AFRY source sha256 must contain 64 lowercase hex digits")
        publisher_release_date = _parse_date(
            item["publisher_release_date"], field="publisher_release_date"
        )
        if publisher_release_date > observed_date:
            raise AfryDataContractError("publisher_release_date cannot be after first observation")
        sources.append(
            AfrySource(
                role=str(item["role"]),
                path=source_path,
                sha256=expected_hash,
                publisher_release_date=publisher_release_date.isoformat(),
                declared_money_basis=str(item.get("declared_money_basis", "")).strip(),
            )
        )
        if not sources[-1].declared_money_basis:
            raise AfryDataContractError("AFRY source declared_money_basis must be explicit")
    roles = [source.role for source in sources]
    if sorted(roles) != ["annual_scenarios", "hourly_prices"]:
        raise AfryDataContractError("registration must contain one annual and one hourly source")
    supporting_documents: list[AfrySupportingDocument] = []
    for item in payload.get("supporting_documents", []):
        relative = Path(str(item["path"]))
        if relative.is_absolute() or relative.suffix.lower() != ".pdf":
            raise AfryDataContractError(
                "AFRY supporting-document paths must be relative .pdf paths"
            )
        document_path = (workspace / relative).resolve()
        try:
            document_path.relative_to(workspace.resolve())
        except ValueError as exc:
            raise AfryDataContractError(
                "AFRY supporting document must be inside the canonical workspace"
            ) from exc
        expected_hash = str(item["sha256"]).lower()
        if re.fullmatch(r"[0-9a-f]{64}", expected_hash) is None:
            raise AfryDataContractError(
                "AFRY supporting-document sha256 must contain 64 lowercase hex digits"
            )
        publisher_release_date = _parse_date(
            item["publisher_release_date"], field="supporting_document.publisher_release_date"
        )
        if publisher_release_date > observed_date:
            raise AfryDataContractError(
                "supporting-document publisher_release_date cannot be after first observation"
            )
        if item.get("content_use") != "semantic_context_only":
            raise AfryDataContractError(
                "AFRY supporting documents must remain semantic_context_only"
            )
        supporting_documents.append(
            AfrySupportingDocument(
                role=str(item["role"]),
                path=document_path,
                sha256=expected_hash,
                publisher_release_date=publisher_release_date.isoformat(),
            )
        )
    document_roles = [document.role for document in supporting_documents]
    if sorted(document_roles) != ["commodity_and_modelling_annex", "quarterly_update_note"]:
        raise AfryDataContractError(
            "registration must contain the AFRY QUN and commodity/modelling annex"
        )
    semantic_item = payload.get("semantic_contract")
    if not isinstance(semantic_item, dict):
        raise AfryDataContractError("registration must contain an AFRY semantic contract")
    semantic_relative = Path(str(semantic_item.get("path", "")))
    if semantic_relative.is_absolute() or semantic_relative.suffix.lower() != ".json":
        raise AfryDataContractError("AFRY semantic-contract path must be a relative .json path")
    semantic_path = (workspace / semantic_relative).resolve()
    try:
        semantic_path.relative_to(workspace.resolve())
    except ValueError as exc:
        raise AfryDataContractError(
            "AFRY semantic contract must be inside the canonical workspace"
        ) from exc
    semantic_hash = str(semantic_item.get("sha256", "")).lower()
    if re.fullmatch(r"[0-9a-f]{64}", semantic_hash) is None:
        raise AfryDataContractError(
            "AFRY semantic-contract sha256 must contain 64 lowercase hex digits"
        )
    return AfryRegistration(
        release_id=release_id,
        first_observed_at_utc=observed_at,
        cadence=cadence,
        expected_arrival_month=expected_arrival_month,
        schedule_timezone=schedule_timezone,
        sources=tuple(sources),
        supporting_documents=tuple(supporting_documents),
        semantic_contract_path=semantic_path,
        semantic_contract_sha256=semantic_hash,
        hourly_contract=hourly_contract,
        annual_contract=annual_contract,
    )


def verify_source(source: AfrySource) -> dict[str, object]:
    if not source.path.is_file():
        raise FileNotFoundError(source.path)
    return _verify_source_payload(source, source.path.read_bytes())


def _verify_source_payload(source: AfrySource, payload: bytes) -> dict[str, object]:
    actual = hashlib.sha256(payload).hexdigest()
    if actual != source.sha256:
        raise AfryDataContractError(
            f"AFRY {source.role} sha256 mismatch: expected={source.sha256} actual={actual}"
        )
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        infos = archive.infolist()
        unsafe = []
        encrypted = []
        for info in infos:
            normalized_name = info.filename.replace("\\", "/")
            member = PurePosixPath(normalized_name)
            if member.is_absolute() or ".." in member.parts or "\x00" in normalized_name:
                unsafe.append(info.filename)
            if info.flag_bits & 0x1:
                encrypted.append(info.filename)
        macro_members = [i.filename for i in infos if i.filename.lower().endswith("vbaproject.bin")]
        active_content_prefixes = (
            "xl/activex/",
            "xl/embeddings/",
            "customui/",
            "xl/macrosheets/",
            "xl/dialogsheets/",
        )
        active_content = [
            i.filename
            for i in infos
            if i.filename.replace("\\", "/").lower().startswith(active_content_prefixes)
        ]
        external_links = [
            info.filename
            for info in infos
            if info.filename.replace("\\", "/").lower().startswith("xl/externallinks/")
        ]
        total_uncompressed = sum(i.file_size for i in infos)
        total_compressed = sum(i.compress_size for i in infos)
        compression_ratio = total_uncompressed / max(total_compressed, 1)
        oversized_members = [
            info.filename
            for info in infos
            if info.file_size > MAX_XLSX_MEMBER_UNCOMPRESSED_BYTES
        ]
        excessive_ratio_members = [
            info.filename
            for info in infos
            if info.file_size / max(info.compress_size, 1) > MAX_XLSX_COMPRESSION_RATIO
        ]
        if unsafe or encrypted or macro_members or active_content:
            raise AfryDataContractError(
                "unsafe workbook container: "
                f"unsafe_paths={len(unsafe)} encrypted={len(encrypted)} "
                f"macros={len(macro_members)} active_content={len(active_content)}"
            )
        if len(infos) > MAX_XLSX_MEMBERS:
            raise AfryDataContractError(f"workbook contains too many ZIP members: {len(infos)}")
        if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise AfryDataContractError(
                f"workbook uncompressed size exceeds contract: {total_uncompressed}"
            )
        if oversized_members:
            raise AfryDataContractError(
                f"workbook member size exceeds contract: members={oversized_members}"
            )
        if excessive_ratio_members:
            raise AfryDataContractError(
                "workbook member compression ratio exceeds contract: "
                f"members={excessive_ratio_members}"
            )
        if compression_ratio > MAX_XLSX_COMPRESSION_RATIO:
            raise AfryDataContractError(
                f"workbook compression ratio exceeds contract: {compression_ratio}"
            )
        formula_members = [
            info.filename
            for info in infos
            if info.filename.replace("\\", "/").lower().startswith("xl/worksheets/")
            and info.filename.lower().endswith(".xml")
            and _zip_member_contains_formula(archive, info)
        ]
    if formula_members:
        raise AfryDataContractError(
            f"workbook worksheet formulas are forbidden: members={formula_members}"
        )
    return {
        "role": source.role,
        "path": source.path.name,
        "sha256": actual,
        "bytes": len(payload),
        "publisher_release_date": source.publisher_release_date,
        "declared_money_basis": source.declared_money_basis,
        "zip_member_count": len(infos),
        "zip_uncompressed_bytes": total_uncompressed,
        "zip_compressed_bytes": total_compressed,
        "zip_compression_ratio": compression_ratio,
        "external_link_member_count": len(external_links),
        "worksheet_formula_member_count": 0,
        "external_links_inert_static_values": bool(external_links),
        "macro_member_count": 0,
        "active_content_member_count": 0,
        "encrypted_member_count": 0,
    }


def verify_supporting_document(document: AfrySupportingDocument) -> dict[str, object]:
    """Verify immutable PDF bytes without executing or rendering the document."""

    if not document.path.is_file():
        raise FileNotFoundError(document.path)
    raw = document.path.read_bytes()
    size = len(raw)
    header = raw[:8]
    if size <= 0 or size > MAX_SUPPORTING_PDF_BYTES:
        raise AfryDataContractError(f"AFRY supporting PDF size violates contract: {size}")
    actual = hashlib.sha256(raw).hexdigest()
    if actual != document.sha256:
        raise AfryDataContractError(
            "AFRY supporting-document sha256 mismatch: "
            f"expected={document.sha256} actual={actual}"
        )
    if not header.startswith(b"%PDF-"):
        raise AfryDataContractError("AFRY supporting document lacks a PDF header")
    forbidden_patterns = {
        "javascript": rb"/JavaScript\b|/JS\b",
        "open_action": rb"/OpenAction\b",
        "additional_action": rb"/AA\b",
        "embedded_file": rb"/EmbeddedFile\b",
        "launch": rb"/Launch\b",
        "rich_media": rb"/RichMedia\b",
    }
    forbidden = [
        name
        for name, pattern in forbidden_patterns.items()
        if re.search(pattern, raw, flags=re.IGNORECASE) is not None
    ]
    return {
        "role": document.role,
        "path": document.path.name,
        "sha256": actual,
        "bytes": size,
        "publisher_release_date": document.publisher_release_date,
        "content_use": "semantic_context_only",
        "parsed_by_model_pipeline": False,
        "active_content_lexical_hints_non_authoritative": forbidden,
        "rendering_or_execution_authorized": False,
    }


def verify_semantic_contract(registration: AfryRegistration) -> dict[str, object]:
    path = registration.semantic_contract_path
    if not path.is_file():
        raise FileNotFoundError(path)
    raw = path.read_bytes()
    actual = hashlib.sha256(raw).hexdigest()
    if actual != registration.semantic_contract_sha256:
        raise AfryDataContractError(
            "AFRY semantic-contract sha256 mismatch: "
            f"expected={registration.semantic_contract_sha256} actual={actual}"
        )
    payload = json.loads(raw.decode("utf-8"))
    if payload.get("schema_version") != "fmv-afry-semantic-contract-v1":
        raise AfryDataContractError("unsupported AFRY semantic-contract schema")
    if payload.get("release_id") != registration.release_id:
        raise AfryDataContractError("AFRY semantic-contract release_id mismatch")
    for field in (
        "contains_vendor_numeric_scenario_values",
        "raw_documents_allowed_in_git",
        "raw_or_derived_numeric_values_allowed_in_git",
        "rag_indexing_allowed",
    ):
        if payload.get(field) is not False:
            raise AfryDataContractError(f"AFRY semantic-contract invariant {field} must be false")
    shaping = payload.get("shaping_admission", {})
    for field in (
        "price_level_authority",
        "monthly_level_authority",
        "scenario_mapping_approved",
        "model_input_authorized",
        "production_authorized",
    ):
        if shaping.get(field) is not False:
            raise AfryDataContractError(
                f"AFRY semantic-contract shaping invariant {field} must be false"
            )
    expected_evidence = {
        (document.role, document.sha256) for document in registration.supporting_documents
    }
    observed_evidence = {
        (str(item.get("role")), str(item.get("sha256")))
        for item in payload.get("source_evidence", [])
    }
    if observed_evidence != expected_evidence:
        raise AfryDataContractError(
            "AFRY semantic-contract document evidence does not match registration"
        )
    policy = payload.get("swiss_policy_crosswalk", {})
    target = policy.get("public_target", {})
    if target.get("delivery_year") != 2035 or target.get("target_twh") != 35.0:
        raise AfryDataContractError("AFRY semantic contract must preserve the Swiss 2035 target")
    if not str(policy.get("afry_reconciliation_status", "")).startswith("NO_GO"):
        raise AfryDataContractError("AFRY policy crosswalk must remain NO_GO while unresolved")
    methodology = payload.get("methodological_evidence")
    if not isinstance(methodology, dict):
        raise AfryDataContractError("AFRY semantic contract missing methodological evidence")
    required_methodology = {
        "evidence_class": "restricted_vendor_model_documentation",
        "observed_swiss_truth": False,
        "independent_validation_evidence": False,
        "probability_calibration_evidence": False,
        "calendar_mapping_authority": False,
        "technical_grid_curtailment_covered": False,
        "negative_price_dynamics_include_behavioral_assumptions": True,
        "hourly_shape_term_structure_documented": True,
    }
    for field, expected in required_methodology.items():
        if methodology.get(field) != expected:
            raise AfryDataContractError(
                f"AFRY methodological invariant {field} must be {expected!r}"
            )
    return {
        "path": path.name,
        "sha256": actual,
        "schema_version": payload["schema_version"],
        "release_id": payload["release_id"],
        "contains_vendor_numeric_scenario_values": False,
        "model_input_authorized": False,
        "production_authorized": False,
    }


def _zip_member_contains_formula(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> bool:
    """Detect worksheet formula elements without materializing large XML members."""

    pattern = re.compile(rb"<(?:[A-Za-z_][\w.-]*:)?f(?:\s|>)", re.IGNORECASE)
    tail = b""
    with archive.open(info) as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            data = tail + chunk
            if pattern.search(data) is not None:
                return True
            tail = data[-64:]
    return False


def _hourly_sheet_contract(
    title: str, contract: AfryHourlyContract
) -> AfryHourlySheetContract:
    matches = [sheet for sheet in contract.sheets if sheet.title == title]
    if len(matches) != 1:
        raise AfryDataContractError(f"hourly sheet is not registered exactly once: {title!r}")
    return matches[0]


def _flush_hourly_batch(writer: pq.ParquetWriter, batch: dict[str, list[object]]) -> None:
    if not batch["delivery_year"]:
        return
    writer.write_table(
        pa.Table.from_pydict(batch, schema=HOURLY_SCHEMA),
        row_group_size=len(batch["delivery_year"]),
    )
    for values in batch.values():
        values.clear()


def materialize_hourly_prices(
    source: Path | io.BytesIO,
    output: Path,
    *,
    release_id: str = RELEASE_ID,
    contract: AfryHourlyContract | None = None,
    declared_money_basis: str = "real January 2026",
) -> dict[str, object]:
    if contract is None:
        raise AfryDataContractError("hourly materialization requires a registered content contract")
    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    batch: dict[str, list[object]] = {field.name: [] for field in HOURLY_SCHEMA}
    writer = pq.ParquetWriter(
        output,
        HOURLY_SCHEMA,
        compression="zstd",
        compression_level=9,
        use_dictionary=True,
        write_statistics=True,
        version="2.6",
    )
    profile: dict[str, object] = {
        "schema_version": "fmv-afry-hourly-profile-v1",
        "release_id": release_id,
        "total_rows": 0,
        "expected_total_rows": contract.expected_total_rows,
        "price_area": contract.price_area,
        "weather_years": list(contract.weather_years),
        "representative_hours_per_group": contract.representative_hours_per_group,
        "declared_money_basis": declared_money_basis,
        "sheets": [],
    }
    seen_sheets: set[str] = set()
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == "ReadMe":
                continue
            sheet_contract = _hourly_sheet_contract(worksheet.title, contract)
            if worksheet.title in seen_sheets:
                raise AfryDataContractError(f"duplicate hourly sheet: {worksheet.title!r}")
            seen_sheets.add(worksheet.title)
            expected_scenario = sheet_contract.scenario_vendor
            expected_years = sheet_contract.delivery_years
            header = tuple(
                cell.value
                for row in worksheet.iter_rows(min_row=13, max_row=13, min_col=2, max_col=7)
                for cell in row
            )
            if header != HOURLY_HEADER:
                raise AfryDataContractError(
                    f"unexpected hourly header in {worksheet.title!r}: {header!r}"
                )
            groups: dict[tuple[int, int], dict[str, object]] = {}
            current_key: tuple[int, int] | None = None
            completed: set[tuple[int, int]] = set()
            last_hour: int | None = None
            empty_rows = 0
            for source_row, row in enumerate(
                worksheet.iter_rows(
                    min_row=14,
                    max_row=worksheet.max_row,
                    min_col=2,
                    max_col=7,
                    values_only=True,
                ),
                start=14,
            ):
                if all(value is None for value in row):
                    empty_rows += 1
                    continue
                if any(value is None for value in row):
                    raise AfryDataContractError(
                        f"partial hourly row in {worksheet.title!r} at Excel row {source_row}"
                    )
                area, scenario, year, weather_year, hour, price = row
                try:
                    year = int(year)
                    weather_year = int(weather_year)
                    hour = int(hour)
                    price = float(price)
                except (TypeError, ValueError) as exc:
                    raise AfryDataContractError(
                        f"invalid hourly types in {worksheet.title!r} row {source_row}"
                    ) from exc
                area = str(area).strip()
                scenario = str(scenario).strip()
                if area != contract.price_area or scenario != expected_scenario:
                    raise AfryDataContractError(
                        f"unexpected area/scenario in {worksheet.title!r} row {source_row}: {area}/{scenario}"
                    )
                if year not in expected_years or weather_year not in contract.weather_years:
                    raise AfryDataContractError(
                        f"unexpected delivery/weather year in {worksheet.title!r} row {source_row}"
                    )
                if not math.isfinite(price):
                    raise AfryDataContractError(
                        f"non-finite price in {worksheet.title!r} row {source_row}"
                    )
                key = (year, weather_year)
                if key != current_key:
                    if current_key is not None:
                        completed.add(current_key)
                    if key in completed:
                        raise AfryDataContractError(
                            f"hourly group reappears in {worksheet.title!r}: {key}"
                        )
                    current_key = key
                    last_hour = None
                if hour != (1 if last_hour is None else last_hour + 1):
                    raise AfryDataContractError(
                        f"non-contiguous representative hour in {worksheet.title!r} {key}: {hour}"
                    )
                last_hour = hour
                group = groups.setdefault(
                    key,
                    {
                        "delivery_year": year,
                        "weather_year": weather_year,
                        "rows": 0,
                        "sum_price": 0.0,
                        "min_price": price,
                        "max_price": price,
                        "strict_negative_prices": 0,
                        "negative_prices": 0,
                        "zero_prices": 0,
                        "hour_of_day_sum": [0.0] * 24,
                        "hour_of_day_count": [0] * 24,
                    },
                )
                group["rows"] = int(group["rows"]) + 1
                group["sum_price"] = float(group["sum_price"]) + price
                group["min_price"] = min(float(group["min_price"]), price)
                group["max_price"] = max(float(group["max_price"]), price)
                group["strict_negative_prices"] = int(group["strict_negative_prices"]) + int(
                    price < 0.0
                )
                group["negative_prices"] = int(group["negative_prices"]) + int(
                    price < -contract.negative_zero_tolerance_eur_mwh
                )
                group["zero_prices"] = int(group["zero_prices"]) + int(price == 0.0)
                hour_of_day = (hour - 1) % 24
                group["hour_of_day_sum"][hour_of_day] += price  # type: ignore[index]
                group["hour_of_day_count"][hour_of_day] += 1  # type: ignore[index]
                values = (
                    release_id,
                    area,
                    scenario,
                    year,
                    weather_year,
                    hour,
                    price,
                    worksheet.title,
                    source_row,
                )
                for field, value in zip(HOURLY_SCHEMA.names, values, strict=True):
                    batch[field].append(value)
                if len(batch["delivery_year"]) == 65_536:
                    _flush_hourly_batch(writer, batch)
            expected_groups = {
                (year, weather) for year in expected_years for weather in contract.weather_years
            }
            if set(groups) != expected_groups:
                missing = sorted(expected_groups - set(groups))
                extra = sorted(set(groups) - expected_groups)
                raise AfryDataContractError(
                    f"hourly group coverage mismatch in {worksheet.title!r}: missing={missing} extra={extra}"
                )
            group_rows = []
            for group in sorted(
                groups.values(), key=lambda item: (item["delivery_year"], item["weather_year"])
            ):
                if group["rows"] != contract.representative_hours_per_group:
                    raise AfryDataContractError(
                        "hourly group has unexpected representative-hour count: "
                        f"{worksheet.title!r} "
                        f"{group['delivery_year']}/{group['weather_year']}={group['rows']}"
                    )
                mean = float(group.pop("sum_price")) / contract.representative_hours_per_group
                hour_means = [
                    total / count
                    for total, count in zip(
                        group.pop("hour_of_day_sum"),  # type: ignore[arg-type]
                        group.pop("hour_of_day_count"),  # type: ignore[arg-type]
                        strict=True,
                    )
                ]
                slot_10_14 = sum(hour_means[10:15]) / 5.0
                slot_18_21 = sum(hour_means[18:22]) / 4.0
                group.update(
                    mean_price=mean,
                    slot_10_14_mean_price=slot_10_14,
                    slot_18_21_mean_price=slot_18_21,
                    slot_18_21_minus_slot_10_14_eur_mwh=slot_18_21 - slot_10_14,
                    slot_10_14_factor=slot_10_14 / mean if mean else None,
                    slot_18_21_factor=slot_18_21 / mean if mean else None,
                    slot_phase_is_swiss_local_time=False,
                    hour_slot_mean_prices_eur_mwh_real_release_basis=hour_means,
                    near_zero_negative_noise_count=(
                        int(group["strict_negative_prices"]) - int(group["negative_prices"])
                    ),
                )
                group_rows.append(group)
            sheet_rows = sum(int(group["rows"]) for group in group_rows)
            profile["total_rows"] = int(profile["total_rows"]) + sheet_rows
            profile["sheets"].append(  # type: ignore[union-attr]
                {
                    "sheet": worksheet.title,
                    "scenario_vendor": expected_scenario,
                    "rows": sheet_rows,
                    "empty_trailing_or_formatted_rows": empty_rows,
                    "groups": group_rows,
                }
            )
        _flush_hourly_batch(writer, batch)
    except Exception:
        writer.close()
        workbook.close()
        output.unlink(missing_ok=True)
        raise
    writer.close()
    workbook.close()
    expected_sheet_titles = {sheet.title for sheet in contract.sheets}
    if seen_sheets != expected_sheet_titles:
        output.unlink(missing_ok=True)
        raise AfryDataContractError(
            "hourly workbook sheet coverage mismatch: "
            f"missing={sorted(expected_sheet_titles - seen_sheets)} "
            f"extra={sorted(seen_sheets - expected_sheet_titles)}"
        )
    if profile["total_rows"] != contract.expected_total_rows:
        output.unlink(missing_ok=True)
        raise AfryDataContractError(f"unexpected total hourly row count: {profile['total_rows']}")
    return profile


def build_hour_slot_shape_benchmark(
    hourly_profile: dict[str, object],
) -> tuple[pd.DataFrame, dict[str, object]]:
    rows: list[dict[str, object]] = []
    group_errors: list[float] = []
    annual_mean_errors: list[float] = []
    release_id = str(hourly_profile["release_id"])
    for sheet in hourly_profile["sheets"]:  # type: ignore[index]
        scenario = str(sheet["scenario_vendor"])
        for group in sheet["groups"]:
            annual_mean = float(group["mean_price"])
            slot_means = list(group["hour_slot_mean_prices_eur_mwh_real_release_basis"])
            if len(slot_means) != 24:
                raise AfryDataContractError("AFRY hour-slot profile must contain exactly 24 slots")
            slot_average = math.fsum(float(value) for value in slot_means) / 24.0
            annual_mean_errors.append(slot_average - annual_mean)
            deviations = [float(value) - slot_average for value in slot_means]
            group_errors.append(math.fsum(deviations) / 24.0)
            for slot, (slot_mean, deviation) in enumerate(
                zip(slot_means, deviations, strict=True)
            ):
                rows.append(
                    {
                        "release_id": release_id,
                        "scenario_vendor": scenario,
                        "delivery_year": int(group["delivery_year"]),
                        "weather_year": int(group["weather_year"]),
                        "representative_hour_slot": slot,
                        "annual_mean_eur_mwh_real_release_basis": annual_mean,
                        "hour_slot_mean_eur_mwh_real_release_basis": float(slot_mean),
                        "annual_zero_mean_additive_shape_eur_mwh": deviation,
                        "multiplicative_factor": (
                            float(slot_mean) / annual_mean if annual_mean != 0.0 else None
                        ),
                        "model_input_authorized": False,
                        "monthly_level_authority": False,
                        "production_authorized": False,
                    }
                )
    max_error = max(abs(error) for error in group_errors)
    max_annual_mean_error = max(abs(error) for error in annual_mean_errors)
    expected_groups = sum(len(sheet["groups"]) for sheet in hourly_profile["sheets"])  # type: ignore[index]
    checks = {
        "group_count_matches_profile": len(group_errors) == expected_groups,
        "twenty_four_slots_per_group": len(rows) == expected_groups * 24,
        "hour_slots_reconcile_to_raw_annual_mean": (
            max_annual_mean_error <= HOUR_SLOT_ANNUAL_MEAN_TOLERANCE_EUR_MWH
        ),
        "annual_additive_shape_zero_mean": (
            max_error <= HOUR_SLOT_ZERO_MEAN_TOLERANCE_EUR_MWH
        ),
    }
    return pd.DataFrame(rows), {
        "schema_version": "fmv-afry-hour-slot-shape-audit-v1",
        "checks": checks,
        "all_checks_pass": all(checks.values()),
        "group_count": len(group_errors),
        "row_count": len(rows),
        "max_abs_group_zero_mean_error_eur_mwh": max_error,
        "zero_mean_tolerance_eur_mwh": HOUR_SLOT_ZERO_MEAN_TOLERANCE_EUR_MWH,
        "max_abs_hour_slots_vs_raw_annual_mean_error_eur_mwh": max_annual_mean_error,
        "annual_mean_tolerance_eur_mwh": HOUR_SLOT_ANNUAL_MEAN_TOLERANCE_EUR_MWH,
        "monthly_zero_mean_guaranteed": False,
        "model_input_authorized": False,
        "monthly_level_authority": False,
        "production_authorized": False,
        "reason": (
            "month, leap-day and DST mapping is not documented; annual hour-slot shape is diagnostic only"
        ),
    }


def _year(value: object) -> int | None:
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return parsed if 2000 <= parsed <= 2100 else None


def _text(values: list[object]) -> str:
    return " | ".join(
        str(value).strip().replace("\n", " ") for value in values if value is not None
    ).strip()


def _scenario_from_text(value: str) -> str | None:
    normalized = value.lower()
    if "delayed transition" in normalized:
        return "Delayed Transition"
    for scenario in ("High", "Central", "Low"):
        if re.search(rf"\b{scenario.lower()}\b", normalized):
            return scenario
    return None


def _value_status(cell: object, year: int) -> str:
    if year <= 2026:
        return "historical"
    font = getattr(cell, "font", None)
    color = getattr(font, "color", None)
    color_type = getattr(color, "type", None)
    if color_type == "rgb" and str(getattr(color, "rgb", "")).upper() in {
        "FF000000",
        "00000000",
    }:
        return "modelled"
    if (
        color_type == "theme"
        and getattr(color, "theme", None) == INTERPOLATED_FONT_THEME
        and math.isclose(
            float(getattr(color, "tint", 0.0)),
            INTERPOLATED_FONT_TINT,
            rel_tol=0.0,
            abs_tol=1e-12,
        )
    ):
        return "interpolated"
    return "unknown"


def extract_annual_facts(
    source: Path | io.BytesIO,
    *,
    release_id: str = RELEASE_ID,
) -> pd.DataFrame:
    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    facts: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        if worksheet.title == "ReadMe":
            continue
        rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row or 1, 1_100),
                min_col=1,
                max_col=min(worksheet.max_column or 1, 100),
            )
        )
        header_rows: list[tuple[int, dict[int, int]]] = []
        for row_number, row in enumerate(rows, start=1):
            years = {cell.column: year for cell in row if (year := _year(cell.value)) is not None}
            if len(years) >= 3:
                header_rows.append((row_number, years))
        for header_index, (header_row, years) in enumerate(header_rows):
            title = (
                _text([cell.value for cell in rows[header_row - 3][1:4]]) if header_row >= 3 else ""
            )
            unit = (
                _text([cell.value for cell in rows[header_row - 2][1:4]]) if header_row >= 2 else ""
            )
            fixed_scenario = _scenario_from_text(title)
            has_next_header = header_index + 1 < len(header_rows)
            next_header = header_rows[header_index + 1][0] if has_next_header else None
            end_row = max(header_row, next_header - 3) if next_header else len(rows)
            current_scenario = fixed_scenario
            filled: list[str | None] = [None, None, None]
            for row_number in range(header_row + 1, end_row + 1):
                row = rows[row_number - 1]
                raw = [
                    str(row[index].value).strip() if row[index].value is not None else None
                    for index in (1, 2, 3)
                ]
                if raw[0] in {"Notes", "Copyright © 2026 AFRY Management Consulting AG"}:
                    break
                row_scenario = next((value for value in raw if value in SCENARIOS), None)
                if fixed_scenario is None and row_scenario is not None:
                    current_scenario = row_scenario
                for index, value in enumerate(raw):
                    if value is not None and value not in SCENARIOS:
                        filled[index] = value
                for column, year in years.items():
                    cell = row[column - 1]
                    value = cell.value
                    if not isinstance(value, (int, float)) or isinstance(value, bool):
                        continue
                    value = float(value)
                    if not math.isfinite(value):
                        raise AfryDataContractError(
                            f"non-finite annual fact in {worksheet.title}!{cell.coordinate}"
                        )
                    value_status = _value_status(cell, year)
                    if value_status == "unknown":
                        raise AfryDataContractError(
                            "unknown AFRY font semantics in "
                            f"{worksheet.title}!{cell.coordinate} for future year {year}"
                        )
                    facts.append(
                        {
                            "release_id": release_id,
                            "sheet": worksheet.title,
                            "table_header_row": header_row,
                            "table_title": title,
                            "unit_raw": unit,
                            "scenario_vendor": current_scenario,
                            "dimension_1_raw": raw[0],
                            "dimension_2_raw": raw[1],
                            "dimension_3_raw": raw[2],
                            "dimension_1": filled[0],
                            "dimension_2": filled[1],
                            "dimension_3": filled[2],
                            "delivery_year": year,
                            "value": value,
                            "value_status": value_status,
                            "source_cell": f"{worksheet.title}!{cell.coordinate}",
                        }
                    )
    workbook.close()
    frame = pd.DataFrame(facts)
    if frame.empty:
        raise AfryDataContractError("annual workbook produced no normalized facts")
    duplicate = frame.duplicated(["source_cell"], keep=False)
    if bool(duplicate.any()):
        raise AfryDataContractError("annual fact extraction produced duplicate source cells")
    return frame.sort_values(
        ["sheet", "table_header_row", "source_cell"], kind="stable"
    ).reset_index(drop=True)


def _one_fact(
    facts: pd.DataFrame,
    *,
    sheet: str,
    title: str,
    scenario: str,
    year: int,
    dimension_1: str | None = None,
    dimension_2: str | None = None,
) -> float | None:
    mask = (
        facts["sheet"].eq(sheet)
        & facts["table_title"].eq(title)
        & facts["scenario_vendor"].eq(scenario)
        & facts["delivery_year"].eq(year)
    )
    if dimension_1 is not None:
        mask &= facts["dimension_1"].eq(dimension_1)
    if dimension_2 is not None:
        mask &= facts["dimension_2"].eq(dimension_2)
    values = facts.loc[mask, "value"]
    if values.empty:
        return None
    if len(values) != 1:
        raise AfryDataContractError(
            "expected one AFRY fact for "
            f"{sheet}/{title}/{scenario}/{year}/{dimension_1}/{dimension_2}, got {len(values)}"
        )
    return float(values.iloc[0])


def _fact_statuses(
    facts: pd.DataFrame,
    *,
    sheet: str,
    title: str,
    scenario: str,
    year: int,
    dimension_1: str | None = None,
    dimension_2: str | None = None,
) -> set[str]:
    if "value_status" not in facts.columns:
        return set()
    mask = (
        facts["sheet"].eq(sheet)
        & facts["table_title"].eq(title)
        & facts["scenario_vendor"].eq(scenario)
        & facts["delivery_year"].eq(year)
    )
    if dimension_1 is not None:
        mask &= facts["dimension_1"].eq(dimension_1)
    if dimension_2 is not None:
        mask &= facts["dimension_2"].eq(dimension_2)
    return set(facts.loc[mask, "value_status"].dropna().astype(str))


def build_structural_benchmark(
    facts: pd.DataFrame,
    *,
    observed_at_utc: str,
    publisher_release_date: str = "2026-06-26",
    release_id: str = RELEASE_ID,
    delivery_years: tuple[int, ...] = tuple(range(2027, 2061)),
) -> pd.DataFrame:
    publication_timestamp = pd.Timestamp(publisher_release_date)
    if publication_timestamp.tz is None:
        publication_timestamp = publication_timestamp.tz_localize("UTC")
    else:
        publication_timestamp = publication_timestamp.tz_convert("UTC")
    observed_timestamp = pd.Timestamp(observed_at_utc)
    if observed_timestamp.tz is None:
        raise AfryDataContractError("observed_at_utc must be timezone-aware")
    observed_timestamp = observed_timestamp.tz_convert("UTC")
    if publication_timestamp > observed_timestamp:
        raise AfryDataContractError("publisher release date cannot be after observation")
    rows: list[dict[str, object]] = []
    capacity_titles = {
        scenario: f"Capacity by Technology - {scenario} scenario" for scenario in SCENARIOS
    }
    generation_titles = {
        scenario: f"Generation by Technology - {scenario} scenario" for scenario in SCENARIOS
    }
    for scenario in SCENARIOS:
        for year in delivery_years:
            cap = lambda technology: _one_fact(  # noqa: E731
                facts,
                sheet="Capacity",
                title=capacity_titles[scenario],
                scenario=scenario,
                year=year,
                dimension_1=technology,
            )
            gen = lambda technology: _one_fact(  # noqa: E731
                facts,
                sheet="Generation",
                title=generation_titles[scenario],
                scenario=scenario,
                year=year,
                dimension_1=technology,
            )
            demand_with = _one_fact(
                facts,
                sheet="Demand",
                title="Annual demand with electrolysis",
                scenario=scenario,
                year=year,
            )
            demand_without = _one_fact(
                facts,
                sheet="Demand",
                title="Annual demand without electrolysis",
                scenario=scenario,
                year=year,
            )
            peak = _one_fact(
                facts,
                sheet="Demand",
                title="Peak demand with electrolysis",
                scenario=scenario,
                year=year,
            )
            pv_gw, pv_twh = cap("Solar PV"), gen("Solar PV")
            wind_gw, wind_twh = cap("Onshore Wind"), gen("Onshore Wind")
            hydro_capacity_parts = [cap("Hydro (Res'voir)"), cap("Hydro (RoR)")]
            hydro_generation_parts = [gen("Hydro (Res'voir)"), gen("Hydro (RoR)")]
            gas_capacity_parts = [cap("CCGT"), cap("CCS Gas")]
            cap_status = lambda technology: _fact_statuses(  # noqa: E731
                facts,
                sheet="Capacity",
                title=capacity_titles[scenario],
                scenario=scenario,
                year=year,
                dimension_1=technology,
            )
            gen_status = lambda technology: _fact_statuses(  # noqa: E731
                facts,
                sheet="Generation",
                title=generation_titles[scenario],
                scenario=scenario,
                year=year,
                dimension_1=technology,
            )
            fact_status = lambda sheet, title: _fact_statuses(  # noqa: E731
                facts,
                sheet=sheet,
                title=title,
                scenario=scenario,
                year=year,
            )
            field_status_sets = {
                "demand_twh": fact_status("Demand", "Annual demand with electrolysis"),
                "peak_load_gw": fact_status("Demand", "Peak demand with electrolysis"),
                "pv_gw": cap_status("Solar PV"),
                "pv_twh": gen_status("Solar PV"),
                "wind_gw": cap_status("Onshore Wind"),
                "wind_twh": gen_status("Onshore Wind"),
                "battery_power_gw": cap_status("Battery"),
                "electrolysis_twh": fact_status(
                    "Demand", "Annual demand with electrolysis"
                )
                | fact_status("Demand", "Annual demand without electrolysis"),
                "hydro_twh": gen_status("Hydro (Res'voir)") | gen_status("Hydro (RoR)"),
                "hydro_capacity_gw": cap_status("Hydro (Res'voir)")
                | cap_status("Hydro (RoR)"),
                "nuclear_gw": cap_status("Nuclear"),
                "gas_gw": cap_status("CCGT") | cap_status("CCS Gas"),
                "net_import_twh": gen_status("Interconnection"),
                "gas_eur_mwh": fact_status(
                    "FuelPrices-Domestic-MWh",
                    "Wholesale gas prices - excluding plant specific delivery costs",
                ),
                "coal_eur_mwh": fact_status(
                    "FuelPrices-Domestic-MWh",
                    "Coal price - excluding plant specific delivery costs",
                ),
                "co2_eur_t": fact_status(
                    "FuelPrices-International", "EU ETS carbon price"
                ),
            }
            field_statuses = {
                field: sorted(statuses) if statuses else ["unknown"]
                for field, statuses in field_status_sets.items()
            }
            status_values = sorted(
                {status for statuses in field_statuses.values() for status in statuses}
            )
            row = {
                "publication_date": publication_timestamp,
                "measurement_date": None,
                "ingested_at_utc": observed_timestamp,
                "source": f"{release_id} restricted client release",
                "scenario": scenario,
                "scenario_edition": release_id,
                "country": "CH",
                "delivery_year": year,
                "delivery_month": None,
                "track": "scenario",
                "quality_flag": "restricted_vendor_benchmark_unmapped",
                "source_value_statuses": "|".join(status_values),
                "field_value_statuses_json": json.dumps(
                    field_statuses,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                ),
                "contains_interpolated_values": "interpolated" in status_values,
                "scenario_weight": None,
                "demand_twh": demand_with,
                "peak_load_gw": peak,
                "pv_gw": pv_gw,
                "pv_twh": pv_twh,
                "wind_gw": wind_gw,
                "wind_twh": wind_twh,
                "battery_power_gw": cap("Battery"),
                "battery_energy_gwh": None,
                "electrolysis_twh": (
                    demand_with - demand_without
                    if demand_with is not None and demand_without is not None
                    else None
                ),
                "hydro_twh": (
                    sum(value for value in hydro_generation_parts if value is not None)
                    if any(value is not None for value in hydro_generation_parts)
                    else None
                ),
                "hydro_capacity_gw": (
                    sum(value for value in hydro_capacity_parts if value is not None)
                    if any(value is not None for value in hydro_capacity_parts)
                    else None
                ),
                "hydro_reservoir_twh": None,
                "nuclear_gw": cap("Nuclear"),
                "gas_gw": (
                    sum(value for value in gas_capacity_parts if value is not None)
                    if any(value is not None for value in gas_capacity_parts)
                    else None
                ),
                "net_import_twh": gen("Interconnection"),
                "gas_eur_mwh": _one_fact(
                    facts,
                    sheet="FuelPrices-Domestic-MWh",
                    title="Wholesale gas prices - excluding plant specific delivery costs",
                    scenario=scenario,
                    year=year,
                ),
                "coal_eur_mwh": _one_fact(
                    facts,
                    sheet="FuelPrices-Domestic-MWh",
                    title="Coal price - excluding plant specific delivery costs",
                    scenario=scenario,
                    year=year,
                ),
                "co2_eur_t": _one_fact(
                    facts,
                    sheet="FuelPrices-International",
                    title="EU ETS carbon price",
                    scenario=scenario,
                    year=year,
                ),
                "model_input_authorized": False,
                "monthly_level_authority": False,
                "production_authorized": False,
            }
            rows.append(row)
    return pd.DataFrame(rows)


def validate_structural_contract(
    benchmark: pd.DataFrame,
    contract: AfryAnnualContract,
) -> dict[str, object]:
    """Require complete structural coverage for every registered scenario/year."""

    expected_keys = {
        (scenario, year) for scenario in SCENARIOS for year in contract.delivery_years
    }
    observed_keys = set(
        zip(benchmark["scenario"], benchmark["delivery_year"], strict=True)
    )
    duplicate_key_count = int(
        benchmark.duplicated(["scenario", "delivery_year"], keep=False).sum()
    )
    null_counts = {
        field: int(benchmark[field].isna().sum())
        for field in contract.structural_required_non_null_fields
    }
    field_status_maps_valid = True
    aggregate_statuses_exact = True
    interpolated_flags_exact = True
    allowed_statuses = {"historical", "modelled", "interpolated"}
    for record in benchmark.to_dict(orient="records"):
        try:
            field_statuses = json.loads(str(record["field_value_statuses_json"]))
        except (KeyError, TypeError, json.JSONDecodeError):
            field_status_maps_valid = False
            aggregate_statuses_exact = False
            interpolated_flags_exact = False
            continue
        if not isinstance(field_statuses, dict) or set(field_statuses) != set(
            contract.structural_required_non_null_fields
        ):
            field_status_maps_valid = False
            continue
        statuses: set[str] = set()
        for values in field_statuses.values():
            if (
                not isinstance(values, list)
                or not values
                or any(value not in allowed_statuses for value in values)
                or values != sorted(set(values))
            ):
                field_status_maps_valid = False
                continue
            statuses.update(values)
        expected_aggregate = "|".join(sorted(statuses))
        if record.get("source_value_statuses") != expected_aggregate:
            aggregate_statuses_exact = False
        if bool(record.get("contains_interpolated_values")) != (
            "interpolated" in statuses
        ):
            interpolated_flags_exact = False
    checks = {
        "scenario_delivery_coverage_exact": observed_keys == expected_keys,
        "scenario_delivery_keys_unique": duplicate_key_count == 0,
        "scenario_delivery_row_count_exact": len(benchmark) == len(expected_keys),
        "required_structural_fields_non_null": all(count == 0 for count in null_counts.values()),
        "declared_nullable_fields_present": all(
            field in benchmark.columns
            for field in contract.structural_declared_nullable_fields
        ),
        "scenario_weights_unassigned": bool(benchmark["scenario_weight"].isna().all()),
        "field_value_status_maps_exact": field_status_maps_valid,
        "aggregate_value_statuses_exact": aggregate_statuses_exact,
        "interpolated_flags_exact": interpolated_flags_exact,
    }
    return {
        "schema_version": "fmv-afry-structural-contract-audit-v2",
        "checks": checks,
        "all_checks_pass": all(checks.values()),
        "expected_rows": len(expected_keys),
        "observed_rows": len(benchmark),
        "duplicate_key_rows": duplicate_key_count,
        "required_field_null_counts": null_counts,
        "delivery_years": list(contract.delivery_years),
        "model_input_authorized": False,
        "monthly_level_authority": False,
        "production_authorized": False,
    }


def validate_annual_semantics(
    facts: pd.DataFrame,
    *,
    delivery_years: tuple[int, ...] = tuple(range(2027, 2061)),
    release_year: int = 2026,
) -> dict[str, object]:
    """Reconcile the annual workbook's critical physical and unit semantics."""

    expected_keys = pd.MultiIndex.from_product(
        [SCENARIOS, delivery_years],
        names=["scenario_vendor", "delivery_year"],
    )
    observed_statuses = set(facts["value_status"].dropna().astype(str))
    allowed_statuses = {"historical", "modelled", "interpolated"}
    observed_future_years = tuple(
        sorted(
            {
                int(year)
                for year in facts.loc[
                    facts["delivery_year"].gt(release_year), "delivery_year"
                ].tolist()
            }
        )
    )
    observed_future_years_by_scenario = {
        scenario: tuple(
            sorted(
                {
                    int(year)
                    for year in facts.loc[
                        facts["scenario_vendor"].eq(scenario)
                        & facts["delivery_year"].gt(release_year),
                        "delivery_year",
                    ].tolist()
                }
            )
        )
        for scenario in SCENARIOS
    }

    generation = (
        facts.loc[facts["sheet"].eq("Generation")]
        .groupby(["scenario_vendor", "delivery_year"], dropna=False)["value"]
        .sum()
        .reindex(expected_keys)
    )
    demand = (
        facts.loc[
            facts["sheet"].eq("Demand")
            & facts["table_title"].eq("Annual demand with electrolysis")
        ]
        .set_index(["scenario_vendor", "delivery_year"])["value"]
        .reindex(expected_keys)
    )
    generation_balance_error = generation - demand

    interconnection = (
        facts.loc[
            facts["sheet"].eq("Generation")
            & facts["dimension_1"].eq("Interconnection")
        ]
        .set_index(["scenario_vendor", "delivery_year"])["value"]
        .reindex(expected_keys)
    )
    net_flows = (
        facts.loc[facts["sheet"].eq("NetFlows")]
        .groupby(["scenario_vendor", "delivery_year"], dropna=False)["value"]
        .sum()
        .reindex(expected_keys)
    )
    net_flow_sign_error = interconnection + net_flows

    pumped_storage = facts.loc[
        facts["sheet"].eq("Generation") & facts["dimension_1"].eq("Pumped Storage"),
        "value",
    ]
    battery = facts.loc[
        facts["sheet"].eq("Generation") & facts["dimension_1"].eq("Battery"),
        "value",
    ]
    unit_expectations = {
        ("Capacity",): "GW",
        ("Generation",): "TWh",
        ("Demand", "Annual demand with electrolysis"): "TWh",
        ("Demand", "Annual demand without electrolysis"): "TWh",
        ("Demand", "Peak demand with electrolysis"): "GW",
        ("Demand", "Peak demand without electrolysis"): "GW",
        (
            "FuelPrices-Domestic-MWh",
            "Wholesale gas prices - excluding plant specific delivery costs",
        ): "€/MWh gross, Real Jan 2026",
        (
            "FuelPrices-Domestic-MWh",
            "Coal price - excluding plant specific delivery costs",
        ): "€/MWh gross, Real Jan 2026",
        ("FuelPrices-International", "EU ETS carbon price"): "€/tCO2, Real Jan 2026",
    }
    unit_mismatches: list[dict[str, object]] = []
    for selector, expected_unit in unit_expectations.items():
        mask = facts["sheet"].eq(selector[0])
        if len(selector) == 2:
            mask &= facts["table_title"].eq(selector[1])
        actual_units = sorted(set(facts.loc[mask, "unit_raw"].astype(str)))
        if actual_units != [expected_unit]:
            unit_mismatches.append(
                {
                    "selector": list(selector),
                    "expected": expected_unit,
                    "actual": actual_units,
                }
            )

    max_generation_error = float(generation_balance_error.abs().max())
    max_net_flow_error = float(net_flow_sign_error.abs().max())
    checks = {
        "observed_future_delivery_year_domain_exact": (
            observed_future_years == delivery_years
        ),
        "scenario_future_delivery_year_domains_exact": all(
            years == delivery_years
            for years in observed_future_years_by_scenario.values()
        ),
        "value_status_vocabulary_exact": observed_statuses <= allowed_statuses
        and bool(observed_statuses),
        "generation_and_demand_key_coverage_exact": bool(
            generation.notna().all() and demand.notna().all()
        ),
        "generation_reconciles_to_demand": (
            math.isfinite(max_generation_error)
            and max_generation_error <= ANNUAL_ENERGY_BALANCE_TOLERANCE_TWH
        ),
        "interconnection_and_net_flow_key_coverage_exact": bool(
            interconnection.notna().all() and net_flows.notna().all()
        ),
        "interconnection_is_negative_sum_of_bilateral_net_flows": (
            math.isfinite(max_net_flow_error)
            and max_net_flow_error <= NET_FLOW_ROUNDING_TOLERANCE_TWH
        ),
        "pumped_storage_generation_is_signed_net_consumption": bool(
            len(pumped_storage) == len(expected_keys) and pumped_storage.le(0.0).all()
        ),
        "battery_generation_is_nonnegative_net_injection": bool(
            len(battery) == len(expected_keys) and battery.ge(0.0).all()
        ),
        "critical_units_exact": not unit_mismatches,
    }
    return {
        "schema_version": "fmv-afry-annual-semantic-profile-v2",
        "checks": checks,
        "all_checks_pass": all(checks.values()),
        "release_year": release_year,
        "registered_future_delivery_years": list(delivery_years),
        "observed_future_delivery_years": list(observed_future_years),
        "observed_future_delivery_years_by_scenario": {
            scenario: list(years)
            for scenario, years in observed_future_years_by_scenario.items()
        },
        "generation_balance_tolerance_twh": ANNUAL_ENERGY_BALANCE_TOLERANCE_TWH,
        "max_abs_generation_minus_demand_twh": max_generation_error,
        "net_flow_rounding_tolerance_twh": NET_FLOW_ROUNDING_TOLERANCE_TWH,
        "max_abs_interconnection_plus_bilateral_net_flows_twh": max_net_flow_error,
        "unit_mismatches": unit_mismatches,
        "semantic_notes": {
            "net_import_twh": "Generation/Interconnection; positive adds supply and equals negative bilateral NetFlows sum",
            "pumped_storage": "signed net consumption in the AFRY Generation table",
            "battery": "nonnegative net injection in the AFRY Generation table",
            "wind": "Onshore Wind only; no offshore category is present for Switzerland",
            "hydro_capacity_gw": "reservoir plus run-of-river; excludes Pumped Storage",
        },
    }


def build_policy_target_gap_profile(
    facts: pd.DataFrame, *, release_id: str = RELEASE_ID
) -> dict[str, object]:
    """Compare AFRY 2035 generation proxies with the Swiss normative target.

    The result is restricted derived data and must remain below ``build/``.  It
    is intentionally not an admission decision and does not map vendor
    scenarios onto FMV scenario probabilities.
    """

    scenario_rows = []
    for scenario in SCENARIOS:
        title = f"Generation by Technology - {scenario} scenario"
        components: dict[str, float] = {}
        for technology in (*AFRY_NON_HYDRO_RENEWABLE_PROXY_TECHNOLOGIES, "Other Thermal"):
            value = _one_fact(
                facts,
                sheet="Generation",
                title=title,
                scenario=scenario,
                year=2035,
                dimension_1=technology,
            )
            if value is None:
                raise AfryDataContractError(
                    f"missing 2035 policy crosswalk component for {scenario}/{technology}"
                )
            components[technology] = value
        strict_proxy = sum(
            components[technology]
            for technology in AFRY_NON_HYDRO_RENEWABLE_PROXY_TECHNOLOGIES
        )
        broad_proxy = strict_proxy + components["Other Thermal"]
        scenario_rows.append(
            {
                "scenario_vendor": scenario,
                "delivery_year": 2035,
                "afry_strict_non_hydro_renewable_proxy_twh": strict_proxy,
                "afry_broad_proxy_including_all_other_thermal_twh": broad_proxy,
                "swiss_legal_target_twh": SWISS_2035_NON_HYDRO_RENEWABLE_TARGET_TWH,
                "strict_proxy_shortfall_twh": (
                    SWISS_2035_NON_HYDRO_RENEWABLE_TARGET_TWH - strict_proxy
                ),
                "broad_proxy_shortfall_twh": (
                    SWISS_2035_NON_HYDRO_RENEWABLE_TARGET_TWH - broad_proxy
                ),
            }
        )
    all_broad_below = all(row["broad_proxy_shortfall_twh"] > 0.0 for row in scenario_rows)
    return {
        "schema_version": "fmv-afry-swiss-policy-gap-profile-v1",
        "release_id": release_id,
        "restricted_derived_values": True,
        "allowed_location": "build_only",
        "target": {
            "delivery_year": 2035,
            "target_twh": SWISS_2035_NON_HYDRO_RENEWABLE_TARGET_TWH,
            "scope": "Swiss domestic electricity production from renewables excluding hydropower",
            "authority": "Swiss Energy Act Article 2 target / SFOE monitoring",
            "source_url": "https://pubdb.bfe.admin.ch/en/publication/download/12427",
        },
        "proxy_definition": {
            "strict_technologies": list(AFRY_NON_HYDRO_RENEWABLE_PROXY_TECHNOLOGIES),
            "broad_sensitivity_addition": "Other Thermal",
            "reason_for_broad_sensitivity": (
                "tests whether an unresolved renewable share inside Other Thermal could close the gap"
            ),
        },
        "scenarios": scenario_rows,
        "all_scenarios_below_target_even_under_broad_proxy": all_broad_below,
        "policy_reconciliation_status": "NO_GO_UNRESOLVED_SEMANTIC_AND_EXECUTION_GAP",
        "interpretation": (
            "The legal target is normative while AFRY is a market/economic scenario set. "
            "The difference may reflect policy execution, but it cannot be labelled purely political "
            "until net/gross, self-consumption, curtailment and technology scopes are reconciled."
        ),
        "unresolved_definitions": [
            "net versus gross production",
            "behind-the-meter photovoltaic self-consumption",
            "economic and forced curtailment",
            "renewable share of waste, biomass and Other Thermal",
            "policy cut-off and assumed compliance path",
        ],
        "model_input_authorized": False,
        "monthly_level_authority": False,
        "production_authorized": False,
    }


def _hourly_groups(profile: dict[str, object]) -> dict[tuple[str, int, int], dict[str, object]]:
    groups = {}
    for sheet in profile["sheets"]:  # type: ignore[index]
        scenario = str(sheet["scenario_vendor"])
        for group in sheet["groups"]:
            groups[(scenario, int(group["delivery_year"]), int(group["weather_year"]))] = group
    return groups


def reconcile_hourly_to_annual(
    facts: pd.DataFrame,
    hourly_profile: dict[str, object],
    *,
    negative_zero_tolerance_eur_mwh: float = PUBLISHED_NEGATIVE_ZERO_TOLERANCE_EUR_MWH,
) -> dict[str, object]:
    groups = _hourly_groups(hourly_profile)
    errors = []
    negative_errors = []
    for (scenario, year, weather_year), group in groups.items():
        title = f"Wholesale electricity prices - {scenario} Scenario"
        expected = _one_fact(
            facts,
            sheet="Price Weather Patterns",
            title=title,
            scenario=scenario,
            year=year,
            dimension_1=str(weather_year),
        )
        if expected is None:
            raise AfryDataContractError(
                f"missing annual weather price for {scenario}/{year}/{weather_year}"
            )
        errors.append(float(group["mean_price"]) - expected)
        negative_expected = _one_fact(
            facts,
            sheet="Negative Prices",
            title="Count of negative price hours",
            scenario=scenario,
            year=year,
            dimension_2=str(weather_year),
        )
        if negative_expected is None:
            raise AfryDataContractError(
                f"missing negative-hour fact for {scenario}/{year}/{weather_year}"
            )
        negative_errors.append(int(group["negative_prices"]) - int(negative_expected))
    max_error = max(abs(value) for value in errors)
    return {
        "hourly_weather_group_count": len(groups),
        "annual_rounding_tolerance_eur_mwh": 0.00501,
        "max_abs_hourly_mean_vs_annual_error_eur_mwh": max_error,
        "hourly_mean_reconciles_with_rounded_annual": max_error <= 0.00501,
        "max_abs_negative_hour_count_error": max(abs(value) for value in negative_errors),
        "negative_hour_counts_reconcile_exactly": all(value == 0 for value in negative_errors),
        "negative_zero_tolerance_eur_mwh": negative_zero_tolerance_eur_mwh,
        "negative_zero_tolerance_status": "versioned_release_contract_and_reconciled_exactly",
    }


def _write_frame(frame: pd.DataFrame, path: Path, schema: pa.Schema) -> None:
    records = [
        {key: None if pd.isna(value) else value for key, value in record.items()}
        for record in frame.to_dict(orient="records")
    ]
    table = pa.Table.from_pylist(records, schema=schema)
    pq.write_table(
        table,
        path,
        compression="zstd",
        compression_level=9,
        use_dictionary=True,
        write_statistics=True,
        version="2.6",
        row_group_size=65_536,
    )


def materialize_afry_catalog(
    *,
    workspace: Path,
    registration_path: Path,
    output_root: Path,
) -> Path:
    workspace = workspace.resolve()
    registration_path = registration_path.resolve()
    output_root = output_root.resolve()
    try:
        registration_path.relative_to(workspace)
        output_root.relative_to((workspace / "build").resolve())
    except ValueError as exc:
        raise AfryDataContractError(
            "registration must be in workspace and output_root below build/"
        ) from exc
    producer_sources = {
        "ingestion_module": Path(__file__).resolve(),
        "materializer_entrypoint": (
            workspace / "scripts" / "materialize_afry_scenario_catalog.py"
        ).resolve(),
    }
    producer_files = []
    for role, path in producer_sources.items():
        try:
            relative = path.relative_to(workspace)
        except ValueError as exc:
            raise AfryDataContractError(
                f"AFRY producer source must be inside workspace: {path}"
            ) from exc
        if not path.is_file() or path.is_symlink():
            raise AfryDataContractError(f"AFRY producer source is not a plain file: {path}")
        producer_files.append(
            {"role": role, "path": relative.as_posix(), "sha256": file_sha256(path)}
        )
    runtime_executable = Path(sys.executable).resolve()
    try:
        runtime_executable_relative = runtime_executable.relative_to(workspace)
    except ValueError as exc:
        raise AfryDataContractError(
            f"AFRY runtime executable must be inside workspace: {runtime_executable}"
        ) from exc
    runtime_executable_sha256 = file_sha256(runtime_executable)
    runtime_prefix = runtime_executable.parent.resolve()
    runtime_prefix_relative = runtime_prefix.relative_to(workspace)
    if output_root == runtime_prefix or output_root.is_relative_to(
        runtime_prefix
    ) or runtime_prefix.is_relative_to(output_root):
        raise AfryDataContractError(
            "AFRY runtime prefix and catalog output root must not overlap"
        )
    runtime_closure = runtime_tree_receipt(runtime_prefix)
    registration_raw = registration_path.read_bytes()
    registration_sha256 = hashlib.sha256(registration_raw).hexdigest()
    registration = load_registration(
        registration_path,
        workspace=workspace,
        raw_bytes=registration_raw,
    )
    registration_receipt = {
        "path": registration_path.relative_to(workspace).as_posix(),
        "schema_version": REGISTRATION_SCHEMA,
        "sha256": registration_sha256,
        "bytes": len(registration_raw),
        "stable_during_materialization": True,
    }
    observed_at = registration.first_observed_at_utc
    supporting_document_receipts = [
        verify_supporting_document(document) for document in registration.supporting_documents
    ]
    semantic_contract_receipt = verify_semantic_contract(registration)
    output_root.mkdir(parents=True, exist_ok=True)
    staging = output_root / f".staging-{uuid.uuid4().hex}"
    staging.mkdir()
    try:
        captures = [capture_verified_source(source) for source in registration.sources]
        captured_sources = [capture[0] for capture in captures]
        source_receipts = [capture[1] for capture in captures]
        by_role = {source.role: source for source in captured_sources}
        hourly_path = staging / "hourly_prices.parquet"
        profile = materialize_hourly_prices(
            io.BytesIO(by_role["hourly_prices"].payload),
            hourly_path,
            release_id=registration.release_id,
            contract=registration.hourly_contract,
            declared_money_basis=by_role["hourly_prices"].declared_money_basis,
        )
        annual_facts = extract_annual_facts(
            io.BytesIO(by_role["annual_scenarios"].payload),
            release_id=registration.release_id,
        )
        annual_semantics = validate_annual_semantics(
            annual_facts,
            delivery_years=registration.annual_contract.delivery_years,
            release_year=registration.annual_contract.release_year,
        )
        policy_gap = build_policy_target_gap_profile(
            annual_facts, release_id=registration.release_id
        )
        annual_path = staging / "annual_facts.parquet"
        _write_frame(annual_facts, annual_path, ANNUAL_FACT_SCHEMA)
        structural = build_structural_benchmark(
            annual_facts,
            observed_at_utc=observed_at,
            publisher_release_date=by_role["annual_scenarios"].publisher_release_date,
            release_id=registration.release_id,
            delivery_years=registration.annual_contract.delivery_years,
        )
        structural_contract_audit = validate_structural_contract(
            structural,
            registration.annual_contract,
        )
        structural_path = staging / "structural_scenarios_benchmark.parquet"
        _write_frame(structural, structural_path, STRUCTURAL_SCHEMA)
        hour_slot_shape, hour_slot_shape_audit = build_hour_slot_shape_benchmark(profile)
        hour_slot_shape_path = staging / "hour_slot_shape_benchmark.parquet"
        _write_frame(hour_slot_shape, hour_slot_shape_path, HOUR_SLOT_SHAPE_SCHEMA)
        reconciliation = reconcile_hourly_to_annual(
            annual_facts,
            profile,
            negative_zero_tolerance_eur_mwh=(
                registration.hourly_contract.negative_zero_tolerance_eur_mwh
            ),
        )
        profile_path = staging / "hourly_profile.json"
        profile_path.write_bytes(canonical_json_bytes(profile))
        annual_semantics_path = staging / "annual_semantic_profile.json"
        annual_semantics_path.write_bytes(canonical_json_bytes(annual_semantics))
        structural_contract_path = staging / "structural_contract_audit.json"
        structural_contract_path.write_bytes(canonical_json_bytes(structural_contract_audit))
        policy_gap_path = staging / "swiss_2035_policy_gap_profile.json"
        policy_gap_path.write_bytes(canonical_json_bytes(policy_gap))
        hour_slot_shape_audit_path = staging / "hour_slot_shape_audit.json"
        hour_slot_shape_audit_path.write_bytes(canonical_json_bytes(hour_slot_shape_audit))
        for captured, receipt in zip(captured_sources, source_receipts, strict=True):
            post_hash = hashlib.sha256(captured.payload).hexdigest()
            if post_hash != captured.sha256:
                raise AfryDataContractError(
                    f"AFRY {captured.role} capture changed during materialization"
                )
            receipt["post_materialization_sha256"] = post_hash
            receipt["stable_capture_verified"] = True
        artifacts = []
        for path, rows in (
            (hourly_path, int(profile["total_rows"])),
            (annual_path, len(annual_facts)),
            (structural_path, len(structural)),
            (profile_path, None),
            (annual_semantics_path, None),
            (structural_contract_path, None),
            (policy_gap_path, None),
            (hour_slot_shape_path, len(hour_slot_shape)),
            (hour_slot_shape_audit_path, None),
        ):
            artifacts.append(
                {
                    "name": path.name,
                    "sha256": file_sha256(path),
                    "bytes": path.stat().st_size,
                    "rows": rows,
                }
            )
        gates = {
            "source_hashes_exact": "PASS",
            "source_capture_stable": "PASS",
            "workbook_containers_safe": "PASS",
            "semantic_contract_exact": "PASS",
            "hourly_schema_exact": "PASS",
            "hourly_group_coverage_exact": "PASS",
            "representative_hour_continuity": "PASS",
            "annual_semantic_reconciliation": (
                "PASS" if annual_semantics["all_checks_pass"] else "NO_GO"
            ),
            "structural_contract_complete": (
                "PASS" if structural_contract_audit["all_checks_pass"] else "NO_GO"
            ),
            "swiss_2035_policy_target_reconciliation": policy_gap[
                "policy_reconciliation_status"
            ],
            "annual_hour_slot_shape_zero_mean": (
                "PASS" if hour_slot_shape_audit["all_checks_pass"] else "NO_GO"
            ),
            "hourly_annual_rounding_reconciliation": (
                "PASS" if reconciliation["hourly_mean_reconciles_with_rounded_annual"] else "NO_GO"
            ),
            "negative_hour_count_reconciliation": (
                "PASS" if reconciliation["negative_hour_counts_reconcile_exactly"] else "NO_GO"
            ),
            "model_input_authorized": "NO_GO",
            "production_authorized": "NO_GO",
        }
        for item in producer_files:
            path = workspace / str(item["path"])
            if file_sha256(path) != item["sha256"]:
                raise AfryDataContractError(
                    f"AFRY producer source changed during materialization: {item['path']}"
                )
        if registration_path.read_bytes() != registration_raw:
            raise AfryDataContractError("AFRY registration changed during materialization")
        if file_sha256(runtime_executable) != runtime_executable_sha256:
            raise AfryDataContractError("AFRY runtime executable changed during materialization")
        if runtime_tree_receipt(runtime_prefix) != runtime_closure:
            raise AfryDataContractError("AFRY runtime closure changed during materialization")
        catalog_body = {
            "schema_version": CATALOG_SCHEMA,
            "release_id": registration.release_id,
            "decision_state": CATALOG_DECISION_STATE,
            "first_observed_at_utc": observed_at,
            "registration": registration_receipt,
            "annual_contract": {
                "release_year": registration.annual_contract.release_year,
                "first_delivery_year_offset": (
                    registration.annual_contract.first_delivery_year_offset
                ),
                "delivery_year_count": registration.annual_contract.delivery_year_count,
                "delivery_years": list(registration.annual_contract.delivery_years),
                "scenarios": list(SCENARIOS),
            },
            "sources": source_receipts,
            "supporting_documents": supporting_document_receipts,
            "semantic_contract": semantic_contract_receipt,
            "artifacts": artifacts,
            "gates": gates,
            "integrity_verdict": CATALOG_INTEGRITY_PASS
            if all(gates[name] == "PASS" for name in CATALOG_INTEGRITY_GATE_NAMES)
            else CATALOG_INTEGRITY_NO_GO,
            "producer": {
                "files": producer_files,
                "runtime": {
                    "python": ".".join(str(value) for value in sys.version_info[:3]),
                    "implementation": sys.implementation.name,
                    "cache_tag": str(sys.implementation.cache_tag),
                    "platform_system": platform.system(),
                    "platform_machine": platform.machine(),
                    "executable_path": runtime_executable_relative.as_posix(),
                    "executable_sha256": runtime_executable_sha256,
                    "prefix_path": runtime_prefix_relative.as_posix(),
                    "dependency_closure": runtime_closure,
                    "pandas": pd.__version__,
                    "pyarrow": pa.__version__,
                    "openpyxl": openpyxl.__version__,
                },
                "source_stable_during_materialization": True,
            },
            "admission": {
                "license_classification": "restricted_client_engagement",
                "raw_or_derived_values_allowed_in_git": False,
                "rag_indexing_allowed": False,
                "monthly_level_authority": False,
                "shaping_benchmark_only": True,
                "scenario_mapping_approved": False,
                "model_input_authorized": False,
                "production_authorized": False,
            },
            "time_contract": {
                "hourly_key": [
                    "scenario_vendor",
                    "delivery_year",
                    "weather_year",
                    "representative_hour",
                ],
                "representative_hour_range": [1, 8760],
                "utc_timestamp_available": False,
                "dst_or_leap_mapping_admitted": False,
                "reason": "AFRY supplies 8760 representative hours for every weather year, including 2012",
            },
            "reconciliation": reconciliation,
            "annual_semantics": annual_semantics,
            "datasets": {
                "hourly_prices": {
                    "artifact": "hourly_prices.parquet",
                    "primary_key": [
                        "scenario_vendor",
                        "delivery_year",
                        "weather_year",
                        "representative_hour",
                    ],
                    "purpose": "restricted vendor benchmark and shaping diagnostics",
                },
                "annual_facts": {
                    "artifact": "annual_facts.parquet",
                    "primary_key": ["source_cell"],
                    "purpose": "loss-minimizing normalized access with exact cell provenance",
                },
                "annual_semantic_profile": {
                    "artifact": "annual_semantic_profile.json",
                    "purpose": "unit, sign and annual energy-balance reconciliation",
                },
                "structural_contract_audit": {
                    "artifact": "structural_contract_audit.json",
                    "purpose": "registered scenario/year coverage and required-field completeness",
                },
                "swiss_2035_policy_gap_profile": {
                    "artifact": "swiss_2035_policy_gap_profile.json",
                    "purpose": "restricted local comparison of AFRY proxies with the public Swiss target",
                    "admission": "NO_GO until semantic and policy-execution scopes reconcile",
                },
                "hour_slot_shape_benchmark": {
                    "artifact": "hour_slot_shape_benchmark.parquet",
                    "primary_key": [
                        "scenario_vendor",
                        "delivery_year",
                        "weather_year",
                        "representative_hour_slot"
                    ],
                    "purpose": "annual zero-mean additive hour-slot diagnostic; not month mapped",
                    "admission": "benchmark only; model input NO_GO",
                },
                "hour_slot_shape_audit": {
                    "artifact": "hour_slot_shape_audit.json",
                    "purpose": "zero-mean and coverage evidence for the compact shape benchmark",
                },
                "structural_scenarios_benchmark": {
                    "artifact": "structural_scenarios_benchmark.parquet",
                    "primary_key": ["scenario", "country", "delivery_year"],
                    "purpose": "partial CH structural benchmark; missing fields remain null",
                },
            },
            "agent_access": {
                "recommended_engine": "pyarrow",
                "recommended_engine_reason": "present in the governed local runtime and reads Parquet without a database copy",
                "optional_engine": "DuckDB only when provisioned in a governed runtime",
                "example_python": [
                    "import pyarrow as pa",
                    "import pyarrow.parquet as pq",
                    "from pfc_shaping.data.afry_scenarios import read_verified_artifact_bytes",
                    "catalog, shape_bytes = read_verified_artifact_bytes(bundle, 'hour_slot_shape_benchmark.parquet')",
                    "shape = pq.read_table(pa.BufferReader(shape_bytes), filters=[('scenario_vendor', '=', 'Central')])",
                ],
                "prohibited": [
                    "embedding raw numeric rows",
                    "treating vendor scenarios as probabilities",
                    "using AFRY price levels in place of the monthly EEX solver",
                    "constructing UTC timestamps without a separately governed mapping",
                ],
            },
        }
        bundle_id = catalog_content_id(catalog_body)
        catalog = {**catalog_body, "bundle_id": bundle_id}
        manifest_path = staging / "catalog.json"
        manifest_path.write_bytes(canonical_json_bytes(catalog))
        (staging / "catalog.sha256").write_bytes(
            f"{file_sha256(manifest_path)}\n".encode("ascii")
        )
        final = output_root / bundle_id
        if final.exists():
            existing = final / "catalog.json"
            verify_catalog_bundle(final)
            if _stable_regular_file_bytes(existing) != _stable_regular_file_bytes(
                manifest_path
            ):
                raise AfryDataContractError(
                    f"bundle id collision or corrupt existing bundle: {final}"
                )
            shutil.rmtree(staging)
            return final
        try:
            staging.replace(final)
        except OSError:
            if not final.exists():
                raise
            existing = final / "catalog.json"
            verify_catalog_bundle(final)
            if _stable_regular_file_bytes(existing) != _stable_regular_file_bytes(
                manifest_path
            ):
                raise AfryDataContractError(
                    f"bundle id collision or divergent concurrent bundle: {final}"
                )
            shutil.rmtree(staging)
            return final
        try:
            verify_catalog_bundle(final)
        except Exception:
            shutil.rmtree(final, ignore_errors=True)
            raise
        return final
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise


__all__ = [
    "AfryDataContractError",
    "AfryAnnualContract",
    "AfryCapturedSource",
    "AfryRegistration",
    "AfryHourlyContract",
    "AfryHourlySheetContract",
    "AfrySource",
    "AfrySupportingDocument",
    "CATALOG_SCHEMA",
    "catalog_content_id",
    "capture_verified_source",
    "REGISTRATION_SCHEMA",
    "RELEASE_ID",
    "build_structural_benchmark",
    "build_policy_target_gap_profile",
    "build_hour_slot_shape_benchmark",
    "extract_annual_facts",
    "STRUCTURAL_SCHEMA",
    "STRUCTURAL_DECLARED_NULLABLE_FIELDS",
    "STRUCTURAL_REQUIRED_NON_NULL_FIELDS",
    "HOUR_SLOT_SHAPE_SCHEMA",
    "file_sha256",
    "load_registration",
    "materialize_afry_catalog",
    "materialize_hourly_prices",
    "read_verified_artifact_bytes",
    "reconcile_hourly_to_annual",
    "validate_annual_semantics",
    "validate_structural_contract",
    "verify_catalog_bundle",
    "verify_source",
    "verify_supporting_document",
    "verify_semantic_contract",
    "runtime_tree_receipt",
]
