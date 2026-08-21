"""Strict, benchmark-only intake for frozen OMPEX HFC workbooks.

OMPEX is not truth and it has no training, tuning, selection, monthly-level or
promotion authority.  This module only makes a frozen workbook structurally
comparable with an independently frozen PFC and an independent realised truth.
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
from pathlib import PurePosixPath
from typing import Mapping
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook

SCHEMA_VERSION = "fmv_ompex_benchmark_intake.v1"
EXPECTED_SHEET = "HFC"
EXPECTED_HEADERS = ("Date", "EUR/MWh")
LOCAL_TIMEZONE = "Europe/Zurich"
NATIVE_RESOLUTION = pd.Timedelta(hours=1)
MAX_ZIP_MEMBERS = 64
MAX_MEMBER_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
MAX_TOTAL_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_COMPRESSION_RATIO = 1_000.0

_VINTAGE_NAME = re.compile(
    r"^HFC_Ompex_(?P<date>\d{8})_(?P<time>\d{6})\.xlsx$",
    re.IGNORECASE,
)
_FORBIDDEN_MEMBERS = (
    "xl/vbaproject.bin",
    "xl/connections.xml",
)
_FORBIDDEN_PREFIXES = (
    "xl/externallinks/",
    "xl/embeddings/",
    "customui/",
)


class OmpexBenchmarkError(ValueError):
    """Raised when OMPEX evidence cannot be interpreted unambiguously."""


@dataclass(frozen=True)
class OmpexWorkbook:
    """Parsed hourly curve and a price-free structural receipt."""

    data: pd.DataFrame
    receipt: Mapping[str, object]


def parse_ompex_workbook(
    payload: bytes,
    *,
    source_name: str,
    accept_unconfirmed_hour_end_semantics: bool = False,
) -> OmpexWorkbook:
    """Parse one frozen workbook while failing closed on time semantics.

    The observed workbooks are structurally consistent with local hour-ending
    labels.  That convention is not vendor-authenticated, so callers must opt
    in explicitly and the receipt continues to record the missing authority.
    """

    if not isinstance(payload, bytes) or not payload:
        raise OmpexBenchmarkError("OMPEX payload must be non-empty immutable bytes")
    if not accept_unconfirmed_hour_end_semantics:
        raise OmpexBenchmarkError(
            "OMPEX hour-ending timestamp semantics require explicit acceptance"
        )

    digest = sha256(payload).hexdigest()
    _validate_ooxml_container(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload), read_only=True, data_only=False, keep_links=True
        )
    except Exception as exc:  # pragma: no cover - library exception taxonomy varies
        raise OmpexBenchmarkError("OMPEX workbook cannot be opened") from exc

    try:
        if workbook.sheetnames != [EXPECTED_SHEET]:
            raise OmpexBenchmarkError(
                f"OMPEX workbook must contain only the {EXPECTED_SHEET!r} sheet"
            )
        sheet = workbook[EXPECTED_SHEET]
        if sheet.sheet_state != "visible":
            raise OmpexBenchmarkError("OMPEX HFC sheet must be visible")
        if sheet.max_column != 2:
            raise OmpexBenchmarkError("OMPEX HFC sheet must contain exactly two columns")

        header_cells = tuple(
            sheet.cell(row=1, column=column).value for column in (1, 2)
        )
        if header_cells != EXPECTED_HEADERS:
            raise OmpexBenchmarkError(
                f"OMPEX headers must be exactly {EXPECTED_HEADERS!r}"
            )

        raw_labels: list[object] = []
        prices: list[float] = []
        blank_seen = False
        for date_cell, price_cell in sheet.iter_rows(
            min_row=2, max_col=2, values_only=False
        ):
            if date_cell.data_type == "f" or price_cell.data_type == "f":
                raise OmpexBenchmarkError("OMPEX formulas are forbidden")
            date_value = date_cell.value
            price_value = price_cell.value
            if date_value is None and price_value is None:
                blank_seen = True
                continue
            if blank_seen:
                raise OmpexBenchmarkError("OMPEX data contain an internal blank row")
            if date_value is None or price_value is None:
                raise OmpexBenchmarkError("OMPEX rows require both date and price")
            try:
                price = float(price_value)
            except (TypeError, ValueError) as exc:
                raise OmpexBenchmarkError("OMPEX price is not numeric") from exc
            if not math.isfinite(price):
                raise OmpexBenchmarkError("OMPEX prices must be finite")
            raw_labels.append(date_value)
            prices.append(price)
    finally:
        workbook.close()

    if not raw_labels:
        raise OmpexBenchmarkError("OMPEX workbook has no curve rows")
    local_hour_end = pd.DatetimeIndex([_parse_local_label(value) for value in raw_labels])
    if local_hour_end.tz is not None:
        raise OmpexBenchmarkError("OMPEX source labels must be naive Swiss local time")
    if any(
        value.minute != 0 or value.second != 0 or value.microsecond != 0
        for value in local_hour_end
    ):
        raise OmpexBenchmarkError("OMPEX curve is not native hourly data")

    local_hour_start = local_hour_end - NATIVE_RESOLUTION
    try:
        delivery_start = local_hour_start.tz_localize(
            LOCAL_TIMEZONE, ambiguous="infer", nonexistent="raise"
        ).tz_convert("UTC")
    except Exception as exc:
        raise OmpexBenchmarkError(
            "OMPEX labels do not form an unambiguous Swiss hour-ending curve"
        ) from exc

    if delivery_start.has_duplicates or not delivery_start.is_monotonic_increasing:
        raise OmpexBenchmarkError("OMPEX UTC delivery starts must be unique and sorted")
    if len(delivery_start) > 1:
        cadence = delivery_start[1:] - delivery_start[:-1]
        if not bool((cadence == NATIVE_RESOLUTION).all()):
            raise OmpexBenchmarkError("OMPEX UTC curve must be contiguous hourly data")

    vintage_match = _VINTAGE_NAME.fullmatch(source_name)
    vintage_label = None
    if vintage_match:
        vintage_label = {
            "date_label": vintage_match.group("date"),
            "time_label": vintage_match.group("time"),
            "availability_semantics_authenticated": False,
        }

    data = pd.DataFrame(
        {
            "delivery_start_utc": delivery_start,
            "price_eur_mwh": prices,
        }
    )
    receipt: dict[str, object] = {
        "schema_version": SCHEMA_VERSION,
        "source_name": source_name,
        "source_sha256": digest,
        "source_size_bytes": len(payload),
        "row_count": len(data),
        "native_resolution": "PT1H",
        "native_quarter_hour": False,
        "currency_unit": "EUR/MWh",
        "timestamp_timezone": "UTC",
        "source_label_interpretation": "Europe/Zurich local hour-ending",
        "hour_end_interpretation_structurally_validated": True,
        "hour_end_interpretation_vendor_authenticated": False,
        "filename_vintage_label": vintage_label,
        "price_statistics_emitted": False,
        "authority": {
            "benchmark_only": True,
            "model_input": False,
            "training": False,
            "tuning": False,
            "selection": False,
            "monthly_level": False,
            "promotion": False,
            "production": False,
        },
    }
    return OmpexWorkbook(data=data, receipt=receipt)


def select_post_origin_rows(
    workbook: OmpexWorkbook, *, origin_utc: pd.Timestamp
) -> pd.DataFrame:
    """Return only delivery intervals not started before an authenticated origin."""

    origin = pd.Timestamp(origin_utc)
    if origin.tzinfo is None:
        raise OmpexBenchmarkError("OMPEX scoring origin must be timezone-aware")
    origin = origin.tz_convert("UTC")
    selected = workbook.data.loc[
        workbook.data["delivery_start_utc"].ge(origin)
    ].copy()
    if selected.empty:
        raise OmpexBenchmarkError("OMPEX workbook has no post-origin forecast rows")
    return selected.reset_index(drop=True)


def hourly_transport_to_quarter_hour(frame: pd.DataFrame) -> pd.DataFrame:
    """Stepwise transport for alignment only; never creates native 15-minute truth."""

    required = {"delivery_start_utc", "price_eur_mwh"}
    if set(frame.columns) != required:
        raise OmpexBenchmarkError(
            "OMPEX hourly transport requires only delivery_start_utc and price_eur_mwh"
        )
    rows: list[pd.DataFrame] = []
    for offset in range(4):
        piece = frame.copy()
        piece["delivery_start_utc"] = piece["delivery_start_utc"] + pd.Timedelta(
            minutes=15 * offset
        )
        piece["native_quarter_hour"] = False
        piece["effective_hour_cluster"] = frame["delivery_start_utc"].to_numpy()
        rows.append(piece)
    return (
        pd.concat(rows, ignore_index=True)
        .sort_values("delivery_start_utc")
        .reset_index(drop=True)
    )


def _parse_local_label(value: object) -> pd.Timestamp:
    if isinstance(value, (datetime, date, pd.Timestamp)):
        parsed = pd.Timestamp(value)
    else:
        try:
            parsed = pd.to_datetime(
                str(value).strip(), format="%d.%m.%Y %H:%M", errors="raise"
            )
        except (TypeError, ValueError) as exc:
            raise OmpexBenchmarkError("OMPEX date label is invalid") from exc
    if pd.isna(parsed):
        raise OmpexBenchmarkError("OMPEX date label is invalid")
    return pd.Timestamp(parsed)


def _validate_ooxml_container(payload: bytes) -> None:
    try:
        with ZipFile(BytesIO(payload)) as archive:
            members = archive.infolist()
            if not members or len(members) > MAX_ZIP_MEMBERS:
                raise OmpexBenchmarkError("OMPEX OOXML member count is invalid")
            total_size = 0
            for member in members:
                normalized = member.filename.lower()
                path = PurePosixPath(member.filename)
                if (
                    "\\" in member.filename
                    or path.is_absolute()
                    or ".." in path.parts
                    or ":" in member.filename
                ):
                    raise OmpexBenchmarkError("OMPEX OOXML member path is unsafe")
                if normalized in _FORBIDDEN_MEMBERS or normalized.startswith(
                    _FORBIDDEN_PREFIXES
                ):
                    raise OmpexBenchmarkError("OMPEX active or linked content is forbidden")
                if member.file_size > MAX_MEMBER_UNCOMPRESSED_BYTES:
                    raise OmpexBenchmarkError("OMPEX OOXML member is too large")
                total_size += member.file_size
                if total_size > MAX_TOTAL_UNCOMPRESSED_BYTES:
                    raise OmpexBenchmarkError("OMPEX OOXML payload is too large")
                if member.file_size and member.compress_size == 0:
                    raise OmpexBenchmarkError("OMPEX OOXML compression metadata is invalid")
                if member.compress_size:
                    ratio = member.file_size / member.compress_size
                    if ratio > MAX_COMPRESSION_RATIO:
                        raise OmpexBenchmarkError("OMPEX OOXML compression ratio is unsafe")
    except BadZipFile as exc:
        raise OmpexBenchmarkError("OMPEX payload is not a valid OOXML container") from exc
