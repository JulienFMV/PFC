"""Stable, price-free inventory of a read-only OMPEX workbook archive.

The archive is benchmark evidence only.  Capturing its filenames, bytes and
OOXML structure does not authenticate forecast availability or grant any model
authority.
"""

from __future__ import annotations

import json
import os
import re
import stat
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from hashlib import sha256
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Mapping
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

SCHEMA_VERSION = "fmv_ompex_archive_inventory.v1"
EXPECTED_CURVE_SHEET = "HFC"
EXPECTED_CURVE_HEADERS = ("Date", "EUR/MWh")
MAX_FILE_BYTES = 4 * 1024 * 1024
MAX_ZIP_MEMBERS = 64
MAX_TOTAL_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
_CURVE_NAME = re.compile(
    r"^HFC_Ompex_(?P<date>\d{8})_(?P<time>\d{6})\.xlsx$",
    re.IGNORECASE,
)
_FORMULA_TAG = re.compile(rb"<f(?:\s|>)")
_FORBIDDEN_MEMBERS = {"xl/vbaproject.bin", "xl/connections.xml"}
_FORBIDDEN_PREFIXES = ("xl/externallinks/", "xl/embeddings/", "customui/")


class OmpexArchiveInventoryError(ValueError):
    """Raised when a stable, unambiguous archive inventory cannot be made."""


@dataclass(frozen=True)
class OmpexArchiveInventory:
    """Current exact file inventory and price-free quality summary."""

    files: tuple[Mapping[str, object], ...]
    summary: Mapping[str, object]


def capture_ompex_archive_inventory(
    source_root: Path,
    *,
    previous_observation: Mapping[str, object] | None = None,
) -> OmpexArchiveInventory:
    """Hash every direct XLSX member with stable-read and directory checks."""

    root = Path(source_root)
    _lstat_regular_directory(root)
    before_names, before_directories = _directory_snapshot(root)
    if not before_names:
        raise OmpexArchiveInventoryError("OMPEX archive is empty")

    rows: list[dict[str, object]] = []
    identities: dict[str, tuple[int, int, int, int, int]] = {}
    for name in before_names:
        path = root / name
        payload, identity = _stable_read(path)
        identities[name] = identity
        rows.append(_profile_workbook(name, payload, identity))

    after_names, after_directories = _directory_snapshot(root)
    if after_names != before_names:
        raise OmpexArchiveInventoryError("OMPEX archive member set changed during capture")
    if after_directories != before_directories:
        raise OmpexArchiveInventoryError(
            "OMPEX archive directory tree changed during capture"
        )
    for name, expected_identity in identities.items():
        if _file_identity(os.lstat(root / name)) != expected_identity:
            raise OmpexArchiveInventoryError(
                f"OMPEX source changed after capture: {name}"
            )

    rows.sort(key=lambda row: str(row["file_name"]))
    summary = _summarize(rows, previous_observation=previous_observation)
    return OmpexArchiveInventory(files=tuple(rows), summary=summary)


def _directory_snapshot(
    root: Path,
) -> tuple[tuple[str, ...], dict[str, tuple[int, int, int, int]]]:
    names: list[str] = []
    directories: dict[str, tuple[int, int, int, int]] = {}
    pending = [root]
    while pending:
        directory = pending.pop()
        directory_stat = os.lstat(directory)
        if _is_reparse_or_link(directory, directory_stat) or not stat.S_ISDIR(
            directory_stat.st_mode
        ):
            raise OmpexArchiveInventoryError(
                "OMPEX archive contains a non-regular directory"
            )
        relative_directory = directory.relative_to(root)
        if len(relative_directory.parts) > 3:
            raise OmpexArchiveInventoryError(
                "OMPEX archive directory depth exceeds the governed bound"
            )
        directory_key = (
            "." if not relative_directory.parts else relative_directory.as_posix()
        )
        directories[directory_key] = _directory_identity(directory_stat)
        for member in directory.iterdir():
            member_stat = os.lstat(member)
            if _is_reparse_or_link(member, member_stat):
                raise OmpexArchiveInventoryError(
                    f"OMPEX archive contains a reparse member: {member.name}"
                )
            if stat.S_ISDIR(member_stat.st_mode):
                pending.append(member)
                continue
            if not stat.S_ISREG(member_stat.st_mode):
                raise OmpexArchiveInventoryError(
                    f"OMPEX archive contains a non-regular member: {member.name}"
                )
            if member.suffix.lower() != ".xlsx":
                raise OmpexArchiveInventoryError(
                    f"OMPEX archive contains a non-XLSX member: {member.name}"
                )
            names.append(member.relative_to(root).as_posix())
    if len(names) != len(set(name.casefold() for name in names)):
        raise OmpexArchiveInventoryError(
            "OMPEX archive contains case-insensitive duplicate names"
        )
    return tuple(sorted(names, key=str.casefold)), dict(sorted(directories.items()))


def _stable_read(path: Path) -> tuple[bytes, tuple[int, int, int, int, int]]:
    before = os.lstat(path)
    if _is_reparse_or_link(path, before) or not stat.S_ISREG(before.st_mode):
        raise OmpexArchiveInventoryError(
            f"OMPEX source is not a regular file: {path.name}"
        )
    if before.st_size <= 0 or before.st_size > MAX_FILE_BYTES:
        raise OmpexArchiveInventoryError(
            f"OMPEX source size is outside the governed bound: {path.name}"
        )
    flags = os.O_RDONLY | getattr(os, "O_BINARY", 0)
    descriptor = os.open(path, flags)
    try:
        opened_before = os.fstat(descriptor)
        chunks: list[bytes] = []
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            chunks.append(chunk)
        opened_after = os.fstat(descriptor)
    finally:
        os.close(descriptor)
    after = os.lstat(path)
    expected = _file_identity(before)
    observed = (
        _file_identity(opened_before),
        _file_identity(opened_after),
        _file_identity(after),
    )
    if any(identity != expected for identity in observed):
        raise OmpexArchiveInventoryError(
            f"OMPEX source changed during stable read: {path.name}"
        )
    payload = b"".join(chunks)
    if len(payload) != before.st_size:
        raise OmpexArchiveInventoryError(
            f"OMPEX source read length changed: {path.name}"
        )
    return payload, expected


def _profile_workbook(
    file_name: str,
    payload: bytes,
    identity: tuple[int, int, int, int, int],
) -> dict[str, object]:
    container = _inspect_container(payload)
    match = _CURVE_NAME.fullmatch(file_name)
    role = "curve_vintage" if match else "non_curve_workbook"

    try:
        workbook = load_workbook(
            BytesIO(payload), read_only=True, data_only=False, keep_links=True
        )
    except Exception as exc:  # pragma: no cover - openpyxl taxonomy varies
        raise OmpexArchiveInventoryError(
            f"OMPEX workbook cannot be opened: {file_name}"
        ) from exc
    try:
        sheet_names = list(workbook.sheetnames)
        visible_sheet_count = sum(
            workbook[name].sheet_state == "visible" for name in sheet_names
        )
        if match:
            if sheet_names != [EXPECTED_CURVE_SHEET]:
                raise OmpexArchiveInventoryError(
                    f"OMPEX curve has unexpected sheets: {file_name}"
                )
            sheet = workbook[EXPECTED_CURVE_SHEET]
            headers = tuple(
                sheet.cell(1, column).value for column in range(1, 3)
            )
            if headers != EXPECTED_CURVE_HEADERS or sheet.max_column != 2:
                raise OmpexArchiveInventoryError(
                    f"OMPEX curve schema differs: {file_name}"
                )
            if sheet.max_row < 2:
                raise OmpexArchiveInventoryError(
                    f"OMPEX curve has no delivery rows: {file_name}"
                )
            data_row_count: int | None = int(sheet.max_row - 1)
            header_signature = _canonical_hash(list(headers))
        else:
            data_row_count = None
            header_signature = None
    finally:
        workbook.close()

    modified_at = datetime.fromtimestamp(identity[3] / 1_000_000_000, tz=timezone.utc)
    row: dict[str, object] = {
        "file_name": file_name,
        "role": role,
        "size_bytes": len(payload),
        "sha256": sha256(payload).hexdigest(),
        "source_last_write_time_utc": modified_at.isoformat(),
        "zip_member_count": container["zip_member_count"],
        "zip_member_names_sha256": container["zip_member_names_sha256"],
        "zip_total_uncompressed_bytes": container[
            "zip_total_uncompressed_bytes"
        ],
        "formula_tag_present": container["formula_tag_present"],
        "active_or_linked_content_present": container[
            "active_or_linked_content_present"
        ],
        "sheet_count": len(sheet_names),
        "visible_sheet_count": visible_sheet_count,
        "curve_header_sha256": header_signature,
        "data_row_count": data_row_count,
        "price_values_or_statistics_emitted": False,
    }
    if match:
        valuation_date = datetime.strptime(match.group("date"), "%Y%m%d").date()
        valuation_time = datetime.strptime(match.group("time"), "%H%M%S").time()
        row.update(
            {
                "filename_date": valuation_date.isoformat(),
                "filename_time": valuation_time.isoformat(),
                "filename_availability_authenticated": False,
            }
        )
    else:
        row.update(
            {
                "filename_date": None,
                "filename_time": None,
                "filename_availability_authenticated": False,
            }
        )
    return row


def _inspect_container(payload: bytes) -> dict[str, object]:
    try:
        with ZipFile(BytesIO(payload)) as archive:
            members = archive.infolist()
            if not members or len(members) > MAX_ZIP_MEMBERS:
                raise OmpexArchiveInventoryError(
                    "OMPEX OOXML member count is outside the governed bound"
                )
            names: list[str] = []
            total = 0
            formula_tag_present = False
            active_content = False
            for member in members:
                name = member.filename
                normalized = name.lower()
                path = PurePosixPath(name)
                if (
                    "\\" in name
                    or path.is_absolute()
                    or ".." in path.parts
                    or ":" in name
                ):
                    raise OmpexArchiveInventoryError(
                        "OMPEX OOXML contains an unsafe member path"
                    )
                total += member.file_size
                if total > MAX_TOTAL_UNCOMPRESSED_BYTES:
                    raise OmpexArchiveInventoryError(
                        "OMPEX OOXML uncompressed size is outside the governed bound"
                    )
                if normalized in _FORBIDDEN_MEMBERS or normalized.startswith(
                    _FORBIDDEN_PREFIXES
                ):
                    active_content = True
                if normalized.startswith("xl/worksheets/") and normalized.endswith(
                    ".xml"
                ):
                    formula_tag_present = formula_tag_present or bool(
                        _FORMULA_TAG.search(archive.read(member))
                    )
                names.append(name)
    except BadZipFile as exc:
        raise OmpexArchiveInventoryError(
            "OMPEX source is not a valid OOXML container"
        ) from exc
    return {
        "zip_member_count": len(names),
        "zip_member_names_sha256": _canonical_hash(sorted(names)),
        "zip_total_uncompressed_bytes": total,
        "formula_tag_present": formula_tag_present,
        "active_or_linked_content_present": active_content,
    }


def _summarize(
    rows: list[dict[str, object]],
    *,
    previous_observation: Mapping[str, object] | None,
) -> dict[str, object]:
    curves = [row for row in rows if row["role"] == "curve_vintage"]
    non_curves = [row for row in rows if row["role"] != "curve_vintage"]
    if not curves:
        raise OmpexArchiveInventoryError("OMPEX archive has no dated curve vintages")
    curve_dates = [date.fromisoformat(str(row["filename_date"])) for row in curves]
    if len(curve_dates) != len(set(curve_dates)):
        raise OmpexArchiveInventoryError(
            "OMPEX archive has duplicate filename valuation dates"
        )
    first_date = min(curve_dates)
    last_date = max(curve_dates)
    expected_dates = {
        first_date + timedelta(days=offset)
        for offset in range((last_date - first_date).days + 1)
    }
    observed_dates = set(curve_dates)
    missing_dates = sorted(expected_dates - observed_dates)
    missing_runs = _consecutive_date_runs(missing_dates)
    duplicate_content = [
        {"sha256": digest, "file_names": sorted(names)}
        for digest, names in sorted(
            _group_values(rows, "sha256").items(), key=lambda item: item[0]
        )
        if len(names) > 1
    ]
    row_count_distribution = {
        str(key): value
        for key, value in sorted(
            Counter(int(row["data_row_count"]) for row in curves).items()
        )
    }
    time_distribution = dict(
        sorted(Counter(str(row["filename_time"]) for row in curves).items())
    )
    schema_signatures = Counter(
        (
            str(row["zip_member_names_sha256"]),
            str(row["curve_header_sha256"]),
            int(row["sheet_count"]),
        )
        for row in curves
    )
    previous_names = set()
    previous_file_count = None
    previous_member_set_complete = False
    if previous_observation is not None:
        raw_names = previous_observation.get("file_names", [])
        if not isinstance(raw_names, list) or not all(
            isinstance(value, str) for value in raw_names
        ):
            raise OmpexArchiveInventoryError(
                "previous OMPEX observation file_names must be a string list"
            )
        previous_names = set(raw_names)
        previous_file_count = previous_observation.get("file_count")
        previous_member_set_complete = bool(
            previous_observation.get("member_set_complete", False)
        )
    current_names = {str(row["file_name"]) for row in rows}
    disappeared = sorted(previous_names - current_names)
    appeared = (
        sorted(current_names - previous_names)
        if previous_names and previous_member_set_complete
        else []
    )

    findings: list[dict[str, object]] = []
    if disappeared:
        findings.append(
            {
                "code": "SOURCE_ARCHIVE_MEMBER_DISAPPEARED",
                "severity": "HIGH",
                "confidence": "HIGH",
                "evidence": {
                    "previous_file_count": previous_file_count,
                    "current_file_count": len(rows),
                    "disappeared_file_names": disappeared,
                },
                "impact": (
                    "The shared directory is mutable and cannot by itself prove "
                    "the exact vintage set available at an earlier origin."
                ),
            }
        )
    if missing_dates:
        findings.append(
            {
                "code": "FILENAME_DATE_COVERAGE_GAPS",
                "severity": "HIGH",
                "confidence": "HIGH",
                "evidence": {
                    "missing_date_count": len(missing_dates),
                    "expected_date_count": len(expected_dates),
                    "missing_rate": len(missing_dates) / len(expected_dates),
                    "maximum_consecutive_missing_days": max(
                        run["day_count"] for run in missing_runs
                    ),
                },
                "impact": (
                    "A nominal daily rolling-origin panel would contain "
                    "unexplained origin gaps."
                ),
            }
        )
    findings.extend(
        [
            {
                "code": "FILENAME_AVAILABILITY_NOT_AUTHENTICATED",
                "severity": "CRITICAL",
                "confidence": "HIGH",
                "evidence": {
                    "candidate_filename_origin_count": len(curves),
                    "countable_scientific_origin_count": 0,
                },
                "impact": (
                    "Filename timestamps cannot prove what the desk could have "
                    "used at a forecast origin."
                ),
            },
            {
                "code": "LATEST_VINTAGE_TIME_TRAVEL_FORBIDDEN",
                "severity": "CRITICAL",
                "confidence": "HIGH",
                "evidence": {
                    "exact_vintage_hash_required": True,
                    "historical_revisions_previously_observed": True,
                },
                "impact": (
                    "Using the latest workbook for an earlier origin can leak "
                    "later revisions into the benchmark."
                ),
            },
        ]
    )
    if any(bool(row["formula_tag_present"]) for row in curves):
        findings.append(
            {
                "code": "CURVE_FORMULA_CONTENT_PRESENT",
                "severity": "CRITICAL",
                "confidence": "HIGH",
                "evidence": {
                    "affected_file_count": sum(
                        bool(row["formula_tag_present"]) for row in curves
                    )
                },
                "impact": "Curve values would not be immutable literal evidence.",
            }
        )

    curve_formula_file_count = sum(
        bool(row["formula_tag_present"]) for row in curves
    )
    curve_active_file_count = sum(
        bool(row["active_or_linked_content_present"]) for row in curves
    )
    status = (
        "FAIL_CURRENT_BYTE_INVENTORY_NO_GO_ROLLING_ORIGIN"
        if curve_formula_file_count or curve_active_file_count
        else "PASS_CURRENT_BYTE_INVENTORY_NO_GO_ROLLING_ORIGIN"
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "status": status,
        "file_count": len(rows),
        "curve_vintage_count": len(curves),
        "non_curve_workbook_count": len(non_curves),
        "non_curve_file_names": sorted(str(row["file_name"]) for row in non_curves),
        "first_filename_date": first_date.isoformat(),
        "last_filename_date": last_date.isoformat(),
        "expected_calendar_date_count": len(expected_dates),
        "missing_filename_date_count": len(missing_dates),
        "missing_filename_date_rate": len(missing_dates) / len(expected_dates),
        "missing_filename_dates": [value.isoformat() for value in missing_dates],
        "missing_date_runs": missing_runs,
        "filename_time_distribution": time_distribution,
        "curve_data_row_count_distribution": row_count_distribution,
        "curve_schema_signature_count": len(schema_signatures),
        "curve_formula_file_count": curve_formula_file_count,
        "curve_active_or_linked_content_file_count": curve_active_file_count,
        "duplicate_content_groups": duplicate_content,
        "previous_observation": {
            "provided": previous_observation is not None,
            "previous_file_count": previous_file_count,
            "previous_member_set_complete": previous_member_set_complete,
            "disappeared_file_names": disappeared,
            "appeared_file_names": appeared,
        },
        "origin_readiness": {
            "candidate_filename_origin_count": len(curves),
            "countable_scientific_origin_count": 0,
            "exact_vintage_hash_available": True,
            "availability_at_origin_authenticated": False,
            "rolling_origin_scoring_authorized": False,
        },
        "findings": findings,
        "authority": {
            "current_byte_inventory": True,
            "descriptive_archive_quality": True,
            "scientific_rolling_origin": False,
            "model_input": False,
            "training": False,
            "tuning": False,
            "selection": False,
            "promotion": False,
            "production": False,
        },
        "price_values_or_statistics_emitted": False,
    }


def _consecutive_date_runs(values: list[date]) -> list[dict[str, object]]:
    if not values:
        return []
    runs: list[dict[str, object]] = []
    start = values[0]
    previous = values[0]
    for value in values[1:]:
        if value != previous + timedelta(days=1):
            runs.append(
                {
                    "start_date": start.isoformat(),
                    "end_date": previous.isoformat(),
                    "day_count": (previous - start).days + 1,
                }
            )
            start = value
        previous = value
    runs.append(
        {
            "start_date": start.isoformat(),
            "end_date": previous.isoformat(),
            "day_count": (previous - start).days + 1,
        }
    )
    return runs


def _group_values(
    rows: list[dict[str, object]], key: str
) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {}
    for row in rows:
        grouped.setdefault(str(row[key]), []).append(str(row["file_name"]))
    return grouped


def _canonical_hash(value: object) -> str:
    payload = json.dumps(value, separators=(",", ":"), sort_keys=True).encode(
        "utf-8"
    )
    return sha256(payload).hexdigest()


def _lstat_regular_directory(path: Path) -> os.stat_result:
    try:
        path_stat = os.lstat(path)
    except OSError as exc:
        raise OmpexArchiveInventoryError("OMPEX archive directory is unavailable") from exc
    if _is_reparse_or_link(path, path_stat) or not stat.S_ISDIR(path_stat.st_mode):
        raise OmpexArchiveInventoryError(
            "OMPEX archive root must be a non-reparse directory"
        )
    return path_stat


def _is_reparse_or_link(path: Path, path_stat: os.stat_result) -> bool:
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    attributes = int(getattr(path_stat, "st_file_attributes", 0))
    return path.is_symlink() or bool(attributes & reparse_flag)


def _file_identity(path_stat: os.stat_result) -> tuple[int, int, int, int, int]:
    return (
        int(path_stat.st_dev),
        int(path_stat.st_ino),
        int(path_stat.st_size),
        int(path_stat.st_mtime_ns),
        int(path_stat.st_ctime_ns),
    )


def _directory_identity(path_stat: os.stat_result) -> tuple[int, int, int, int]:
    return (
        int(path_stat.st_dev),
        int(path_stat.st_ino),
        int(path_stat.st_mtime_ns),
        int(path_stat.st_ctime_ns),
    )
