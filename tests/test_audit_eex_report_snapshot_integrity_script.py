from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from scripts.audit_eex_report_snapshot_integrity import audit_snapshot_integrity, main


def test_audit_snapshot_integrity_flags_suspicious_far_horizon_zero_wipe(tmp_path: Path) -> None:
    workbook = tmp_path / "eex.xlsx"
    _build_workbook(
        workbook,
        rows=[
            ("2026-07-07", {"Y01_2027_BASE": 96.34, "Y01_2028_BASE": 81.31, "Y01_2029_BASE": 73.97, "Y01_2030_BASE": 69.70}),
            ("2026-07-08", {"Y01_2027_BASE": 97.54, "Y01_2028_BASE": 81.41, "Y01_2029_BASE": 0, "Y01_2030_BASE": 0}),
        ],
    )

    summary = audit_snapshot_integrity(workbook=workbook, market="CH")

    assert summary["status"] == "SUSPICIOUS_ZERO_WIPE"
    assert summary["zero_wipe_count"] == 2
    assert summary["latest_far_horizon_nonzero_count"] == 0
    assert summary["previous_far_horizon_nonzero_count"] == 2


def test_audit_snapshot_integrity_passes_when_far_horizon_quotes_remain_present(tmp_path: Path) -> None:
    workbook = tmp_path / "eex.xlsx"
    _build_workbook(
        workbook,
        rows=[
            ("2026-07-07", {"Y01_2027_BASE": 96.34, "Y01_2028_BASE": 81.31, "Y01_2029_BASE": 73.97, "Y01_2030_BASE": 69.70}),
            ("2026-07-08", {"Y01_2027_BASE": 97.54, "Y01_2028_BASE": 81.41, "Y01_2029_BASE": 74.10, "Y01_2030_BASE": 69.85}),
        ],
    )

    summary = audit_snapshot_integrity(workbook=workbook, market="CH")

    assert summary["status"] == "PASS"
    assert summary["zero_wipe_count"] == 0


def test_audit_snapshot_integrity_flags_empty_latest_snapshot(tmp_path: Path) -> None:
    workbook = tmp_path / "eex.xlsx"
    _build_workbook(
        workbook,
        rows=[
            ("2026-07-30", {"Y01_2027_BASE": 0, "Y01_2028_BASE": 0, "Y01_2029_BASE": 0, "Y01_2030_BASE": 0}),
            ("2026-07-31", {"Y01_2027_BASE": 0, "Y01_2028_BASE": 0, "Y01_2029_BASE": 0, "Y01_2030_BASE": 0}),
        ],
    )

    summary = audit_snapshot_integrity(workbook=workbook, market="CH")

    assert summary["status"] == "LATEST_SNAPSHOT_EMPTY"
    assert summary["latest_far_horizon_nonzero_count"] == 0
    assert summary["latest_near_horizon_nonzero_count"] == 0


def test_audit_snapshot_integrity_ignores_future_placeholder_rows(tmp_path: Path) -> None:
    workbook = tmp_path / "eex.xlsx"
    _build_workbook(
        workbook,
        rows=[
            ("2026-07-07", {"Y01_2027_BASE": 96.34, "Y01_2028_BASE": 81.31, "Y01_2029_BASE": 73.97, "Y01_2030_BASE": 69.70}),
            ("2026-07-08", {"Y01_2027_BASE": 97.54, "Y01_2028_BASE": 81.41, "Y01_2029_BASE": 74.10, "Y01_2030_BASE": 69.85}),
            ("2026-07-31", {"Y01_2027_BASE": 0, "Y01_2028_BASE": 0, "Y01_2029_BASE": 0, "Y01_2030_BASE": 0}),
        ],
    )

    summary = audit_snapshot_integrity(workbook=workbook, market="CH")

    assert summary["status"] == "PASS"
    assert summary["latest_workbook_date"] == "2026-07-31"
    assert summary["latest_usable_date"] == "2026-07-08"
    assert summary["quote_counts_by_date"]["2026-07-31"] == 0


def test_main_exits_nonzero_on_suspicious_zero_wipe(tmp_path: Path) -> None:
    workbook = tmp_path / "eex.xlsx"
    output = tmp_path / "audit.json"
    _build_workbook(
        workbook,
        rows=[
            ("2026-07-07", {"Y01_2027_BASE": 96.34, "Y01_2028_BASE": 81.31, "Y01_2029_BASE": 73.97}),
            ("2026-07-08", {"Y01_2027_BASE": 97.54, "Y01_2028_BASE": 81.41, "Y01_2029_BASE": 0}),
        ],
    )

    code = main(
        [
            "--workbook",
            str(workbook),
            "--market",
            "CH",
            "--output",
            str(output),
        ]
    )

    assert code == 1
    assert output.exists()


def _build_workbook(path: Path, *, rows: list[tuple[str, dict[str, float]]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CH"
    headers = ["Y01_2027_BASE", "Y01_2028_BASE", "Y01_2029_BASE", "Y01_2030_BASE"]
    for idx, header in enumerate(headers, start=2):
        ws.cell(row=1, column=idx).value = header
    start_row = 4
    for offset, (date_text, values) in enumerate(rows):
        row = start_row + offset
        ws.cell(row=row, column=1).value = date_text
        for idx, header in enumerate(headers, start=2):
            ws.cell(row=row, column=idx).value = values.get(header, 0)
    wb.save(path)
