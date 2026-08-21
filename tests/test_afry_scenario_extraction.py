from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook
from openpyxl.styles import Color, Font

from pfc_shaping.data.afry_scenarios import (
    SCENARIOS,
    STRUCTURAL_DECLARED_NULLABLE_FIELDS,
    STRUCTURAL_REQUIRED_NON_NULL_FIELDS,
    STRUCTURAL_SCHEMA,
    AfryAnnualContract,
    AfryDataContractError,
    _write_frame,
    build_hour_slot_shape_benchmark,
    build_policy_target_gap_profile,
    build_structural_benchmark,
    extract_annual_facts,
    reconcile_hourly_to_annual,
    validate_annual_semantics,
    validate_structural_contract,
)


def _semantic_fixture() -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    def add(
        scenario: str,
        year: int,
        *,
        sheet: str,
        title: str,
        unit: str,
        value: float,
        dimension_1: str | None = None,
    ) -> None:
        rows.append(
            {
                "sheet": sheet,
                "table_title": title,
                "unit_raw": unit,
                "scenario_vendor": scenario,
                "delivery_year": year,
                "dimension_1": dimension_1,
                "value": value,
                "value_status": "modelled",
            }
        )

    for scenario in SCENARIOS:
        for year in range(2027, 2061):
            generation_title = f"Generation by Technology - {scenario} scenario"
            add(
                scenario,
                year,
                sheet="Generation",
                title=generation_title,
                unit="TWh",
                dimension_1="Solar PV",
                value=5.0,
            )
            add(
                scenario,
                year,
                sheet="Generation",
                title=generation_title,
                unit="TWh",
                dimension_1="Pumped Storage",
                value=-1.0,
            )
            add(
                scenario,
                year,
                sheet="Generation",
                title=generation_title,
                unit="TWh",
                dimension_1="Battery",
                value=1.0,
            )
            add(
                scenario,
                year,
                sheet="Generation",
                title=generation_title,
                unit="TWh",
                dimension_1="Interconnection",
                value=5.0,
            )
            for border in ("France", "Germany", "Italy", "Austria"):
                add(
                    scenario,
                    year,
                    sheet="NetFlows",
                    title=f"Net Flows - {scenario} scenario",
                    unit="TWh",
                    dimension_1=f"Switzerland-{border}",
                    value=-1.25,
                )
            for title, unit, value in (
                ("Annual demand with electrolysis", "TWh", 10.0),
                ("Annual demand without electrolysis", "TWh", 9.0),
                ("Peak demand with electrolysis", "GW", 2.0),
                ("Peak demand without electrolysis", "GW", 1.8),
            ):
                add(scenario, year, sheet="Demand", title=title, unit=unit, value=value)
            add(
                scenario,
                year,
                sheet="Capacity",
                title=f"Capacity by Technology - {scenario} scenario",
                unit="GW",
                dimension_1="Solar PV",
                value=10.0,
            )
            for title in (
                "Wholesale gas prices - excluding plant specific delivery costs",
                "Coal price - excluding plant specific delivery costs",
            ):
                add(
                    scenario,
                    year,
                    sheet="FuelPrices-Domestic-MWh",
                    title=title,
                    unit="€/MWh gross, Real Jan 2026",
                    value=50.0,
                )
            add(
                scenario,
                year,
                sheet="FuelPrices-International",
                title="EU ETS carbon price",
                unit="€/tCO2, Real Jan 2026",
                value=100.0,
            )
    return pd.DataFrame(rows)


def test_extract_annual_facts_preserves_cell_provenance_and_status(tmp_path: Path) -> None:
    path = tmp_path / "annual.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Electricity Prices"
    worksheet["B6"] = "Wholesale electricity prices - baseload"
    worksheet["B7"] = "EUR/MWh, Real Jan 2026"
    for column, year in zip((4, 5, 6), (2027, 2028, 2029), strict=True):
        worksheet.cell(row=8, column=column, value=year)
        value = worksheet.cell(row=9, column=column, value=80.0 + column)
        value.font = Font(color="FF000000")
    worksheet["B9"] = "High"
    workbook.save(path)

    facts = extract_annual_facts(path)

    assert facts["source_cell"].tolist() == [
        "Electricity Prices!D9",
        "Electricity Prices!E9",
        "Electricity Prices!F9",
    ]
    assert facts["scenario_vendor"].tolist() == ["High", "High", "High"]
    assert facts["value_status"].tolist() == ["modelled", "modelled", "modelled"]


def test_extract_annual_facts_accepts_only_vendor_grey_as_interpolated(tmp_path: Path) -> None:
    path = tmp_path / "annual-grey.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Electricity Prices"
    worksheet["B6"] = "Wholesale electricity prices - baseload"
    worksheet["B7"] = "EUR/MWh, Real Jan 2026"
    for column, year in zip((4, 5, 6), (2027, 2028, 2029), strict=True):
        worksheet.cell(row=8, column=column, value=year)
        value = worksheet.cell(row=9, column=column, value=80.0 + column)
        value.font = Font(color=Color(theme=0, tint=-0.499984740745262))
    worksheet["B9"] = "Central"
    workbook.save(path)

    facts = extract_annual_facts(path)

    assert facts["value_status"].tolist() == ["interpolated"] * 3


def test_extract_annual_facts_rejects_unknown_future_font_semantics(tmp_path: Path) -> None:
    path = tmp_path / "annual-unknown-font.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Electricity Prices"
    worksheet["B6"] = "Wholesale electricity prices - baseload"
    worksheet["B7"] = "EUR/MWh, Real Jan 2026"
    for column, year in zip((4, 5, 6), (2027, 2028, 2029), strict=True):
        worksheet.cell(row=8, column=column, value=year)
        value = worksheet.cell(row=9, column=column, value=80.0 + column)
        value.font = Font(color="FFFF0000")
    worksheet["B9"] = "High"
    workbook.save(path)

    with pytest.raises(AfryDataContractError, match="unknown AFRY font semantics"):
        extract_annual_facts(path)


def test_reconciliation_uses_separate_weather_dimensions() -> None:
    facts = pd.DataFrame(
        [
            {
                "sheet": "Price Weather Patterns",
                "table_title": "Wholesale electricity prices - Central Scenario",
                "scenario_vendor": "Central",
                "delivery_year": 2030,
                "dimension_1": "2012",
                "dimension_2": None,
                "value": 75.0,
            },
            {
                "sheet": "Negative Prices",
                "table_title": "Count of negative price hours",
                "scenario_vendor": "Central",
                "delivery_year": 2030,
                "dimension_1": None,
                "dimension_2": "2012",
                "value": 2.0,
            },
        ]
    )
    profile = {
        "sheets": [
            {
                "scenario_vendor": "Central",
                "groups": [
                    {
                        "delivery_year": 2030,
                        "weather_year": 2012,
                        "mean_price": 75.004,
                        "negative_prices": 2,
                    }
                ],
            }
        ]
    }

    result = reconcile_hourly_to_annual(facts, profile)

    assert result["hourly_mean_reconciles_with_rounded_annual"] is True
    assert result["negative_hour_counts_reconcile_exactly"] is True
    assert result["negative_zero_tolerance_eur_mwh"] == pytest.approx(5e-6)


def test_hour_slot_shape_is_additive_zero_mean_and_non_authoritative() -> None:
    slot_means = [10.0] * 24
    slot_means[10] = 8.0
    slot_means[20] = 12.0
    profile = {
        "release_id": "AFRY_CH_2026_Q2_V100",
        "sheets": [
            {
                "scenario_vendor": "Central",
                "groups": [
                    {
                        "delivery_year": 2030,
                        "weather_year": 2012,
                        "mean_price": 10.0,
                        "hour_slot_mean_prices_eur_mwh_real_release_basis": slot_means,
                    }
                ],
            }
        ],
    }

    frame, audit = build_hour_slot_shape_benchmark(profile)

    assert len(frame) == 24
    assert frame["annual_zero_mean_additive_shape_eur_mwh"].sum() == pytest.approx(0.0)
    assert not frame["model_input_authorized"].any()
    assert not frame["monthly_level_authority"].any()
    assert audit["all_checks_pass"] is True
    assert audit["monthly_zero_mean_guaranteed"] is False


def test_annual_semantics_reconcile_energy_units_and_signs() -> None:
    result = validate_annual_semantics(_semantic_fixture())

    assert result["all_checks_pass"] is True
    assert result["max_abs_generation_minus_demand_twh"] == pytest.approx(0.0)
    assert result["max_abs_interconnection_plus_bilateral_net_flows_twh"] == pytest.approx(0.0)


def test_annual_semantics_detect_wrong_net_flow_sign() -> None:
    facts = _semantic_fixture()
    facts.loc[facts["sheet"].eq("NetFlows"), "value"] *= -1.0

    result = validate_annual_semantics(facts)

    assert result["all_checks_pass"] is False
    assert result["checks"]["interconnection_is_negative_sum_of_bilateral_net_flows"] is False


def test_annual_semantics_rejects_unregistered_future_year() -> None:
    facts = _semantic_fixture()
    extra = facts.iloc[[0]].copy()
    extra["delivery_year"] = 2061
    facts = pd.concat([facts, extra], ignore_index=True)

    result = validate_annual_semantics(facts, release_year=2026)

    assert result["all_checks_pass"] is False
    assert result["checks"]["observed_future_delivery_year_domain_exact"] is False


def test_policy_target_gap_is_restricted_no_go_even_with_broad_proxy() -> None:
    rows = []
    values = {
        "Solar PV": 20.0,
        "Onshore Wind": 2.0,
        "Other Res": 3.0,
        "CCS Biomass": 0.0,
        "Other Thermal": 2.0,
    }
    for scenario in SCENARIOS:
        for technology, value in values.items():
            rows.append(
                {
                    "sheet": "Generation",
                    "table_title": f"Generation by Technology - {scenario} scenario",
                    "scenario_vendor": scenario,
                    "delivery_year": 2035,
                    "dimension_1": technology,
                    "value": value,
                }
            )

    result = build_policy_target_gap_profile(pd.DataFrame(rows))

    assert result["restricted_derived_values"] is True
    assert result["all_scenarios_below_target_even_under_broad_proxy"] is True
    assert result["policy_reconciliation_status"].startswith("NO_GO")
    assert result["model_input_authorized"] is False
    assert result["scenarios"][0]["strict_proxy_shortfall_twh"] == pytest.approx(10.0)
    assert result["scenarios"][0]["broad_proxy_shortfall_twh"] == pytest.approx(8.0)


def test_structural_benchmark_keeps_missing_fields_null_and_authorities_false() -> None:
    facts = pd.DataFrame(
        columns=[
            "sheet",
            "table_title",
            "scenario_vendor",
            "delivery_year",
            "dimension_1",
            "value",
        ]
    )

    benchmark = build_structural_benchmark(
        facts,
        observed_at_utc="2026-08-03T09:22:51Z",
    )

    assert len(benchmark) == 136
    assert not benchmark["model_input_authorized"].any()
    assert not benchmark["monthly_level_authority"].any()
    assert not benchmark["production_authorized"].any()
    assert benchmark["battery_energy_gwh"].isna().all()


def test_structural_contract_rejects_missing_required_fields() -> None:
    facts = pd.DataFrame(
        columns=[
            "sheet",
            "table_title",
            "scenario_vendor",
            "delivery_year",
            "dimension_1",
            "value",
        ]
    )
    years = (2027,)
    benchmark = build_structural_benchmark(
        facts,
        observed_at_utc="2026-08-03T09:22:51Z",
        delivery_years=years,
    )
    contract = AfryAnnualContract(
        release_year=2026,
        first_delivery_year_offset=1,
        delivery_year_count=len(years),
        delivery_years=years,
        structural_required_non_null_fields=STRUCTURAL_REQUIRED_NON_NULL_FIELDS,
        structural_declared_nullable_fields=STRUCTURAL_DECLARED_NULLABLE_FIELDS,
    )

    audit = validate_structural_contract(benchmark, contract)

    assert audit["all_checks_pass"] is False
    assert audit["checks"]["scenario_delivery_coverage_exact"] is True
    assert audit["checks"]["required_structural_fields_non_null"] is False


def test_structural_contract_accepts_complete_registered_grid() -> None:
    facts = pd.DataFrame(
        columns=[
            "sheet",
            "table_title",
            "scenario_vendor",
            "delivery_year",
            "dimension_1",
            "value",
        ]
    )
    years = (2028, 2029)
    benchmark = build_structural_benchmark(
        facts,
        observed_at_utc="2027-08-03T09:22:51Z",
        publisher_release_date="2027-07-24",
        release_id="AFRY_CH_2027_Q2_V100",
        delivery_years=years,
    )
    for field in STRUCTURAL_REQUIRED_NON_NULL_FIELDS:
        benchmark[field] = 0.0
    benchmark["field_value_statuses_json"] = json.dumps(
        {field: ["modelled"] for field in STRUCTURAL_REQUIRED_NON_NULL_FIELDS},
        sort_keys=True,
        separators=(",", ":"),
    )
    benchmark["source_value_statuses"] = "modelled"
    benchmark["contains_interpolated_values"] = False
    contract = AfryAnnualContract(
        release_year=2027,
        first_delivery_year_offset=1,
        delivery_year_count=len(years),
        delivery_years=years,
        structural_required_non_null_fields=STRUCTURAL_REQUIRED_NON_NULL_FIELDS,
        structural_declared_nullable_fields=STRUCTURAL_DECLARED_NULLABLE_FIELDS,
    )

    audit = validate_structural_contract(benchmark, contract)

    assert audit["all_checks_pass"] is True
    assert audit["expected_rows"] == len(SCENARIOS) * len(years)


def test_structural_contract_rejects_duplicate_key_and_scenario_weight() -> None:
    years = (2027,)
    benchmark = build_structural_benchmark(
        pd.DataFrame(
            columns=[
                "sheet",
                "table_title",
                "scenario_vendor",
                "delivery_year",
                "dimension_1",
                "value",
            ]
        ),
        observed_at_utc="2026-08-03T09:22:51Z",
        delivery_years=years,
    )
    for field in STRUCTURAL_REQUIRED_NON_NULL_FIELDS:
        benchmark[field] = 0.0
    benchmark["field_value_statuses_json"] = json.dumps(
        {field: ["modelled"] for field in STRUCTURAL_REQUIRED_NON_NULL_FIELDS},
        sort_keys=True,
        separators=(",", ":"),
    )
    benchmark["source_value_statuses"] = "modelled"
    benchmark["contains_interpolated_values"] = False
    benchmark.loc[0, "scenario_weight"] = 0.25
    benchmark = pd.concat([benchmark, benchmark.iloc[[0]]], ignore_index=True)
    contract = AfryAnnualContract(
        release_year=2026,
        first_delivery_year_offset=1,
        delivery_year_count=1,
        delivery_years=years,
        structural_required_non_null_fields=STRUCTURAL_REQUIRED_NON_NULL_FIELDS,
        structural_declared_nullable_fields=STRUCTURAL_DECLARED_NULLABLE_FIELDS,
    )

    audit = validate_structural_contract(benchmark, contract)

    assert audit["all_checks_pass"] is False
    assert audit["checks"]["scenario_delivery_keys_unique"] is False
    assert audit["checks"]["scenario_delivery_row_count_exact"] is False
    assert audit["checks"]["scenario_weights_unassigned"] is False


def test_structural_benchmark_provenance_follows_registered_release() -> None:
    facts = pd.DataFrame(
        columns=[
            "sheet",
            "table_title",
            "scenario_vendor",
            "delivery_year",
            "dimension_1",
            "value",
        ]
    )

    benchmark = build_structural_benchmark(
        facts,
        observed_at_utc="2027-07-30T10:00:00Z",
        publisher_release_date="2027-07-24",
        release_id="AFRY_CH_2027_Q2_V100",
        delivery_years=tuple(range(2028, 2062)),
    )

    assert set(benchmark["source"]) == {
        "AFRY_CH_2027_Q2_V100 restricted client release"
    }
    assert set(benchmark["scenario_edition"]) == {"AFRY_CH_2027_Q2_V100"}
    assert set(benchmark["delivery_year"]) == set(range(2028, 2062))


def test_structural_benchmark_writes_exact_stable_parquet_schema(tmp_path: Path) -> None:
    facts = pd.DataFrame(
        columns=[
            "sheet",
            "table_title",
            "scenario_vendor",
            "delivery_year",
            "dimension_1",
            "value",
        ]
    )
    benchmark = build_structural_benchmark(
        facts,
        observed_at_utc="2026-08-03T09:22:51Z",
    )
    output = tmp_path / "structural.parquet"

    _write_frame(benchmark, output, STRUCTURAL_SCHEMA)

    assert pq.read_schema(output).equals(STRUCTURAL_SCHEMA, check_metadata=True)
    stored = pq.read_table(output).to_pandas()
    assert stored["delivery_month"].isna().all()
    assert stored["battery_energy_gwh"].isna().all()
    assert str(stored["publication_date"].dt.tz) == "UTC"
