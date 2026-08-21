from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from pfc_shaping.validation.ompex_archive_inventory import (
    OmpexArchiveInventoryError,
    capture_ompex_archive_inventory,
)


def _write_curve(path: Path, *, formula: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HFC"
    sheet.append(["Date", "EUR/MWh"])
    sheet.append(["01.01.2026 01:00", "=40+1" if formula else 41.0])
    sheet.append(["01.01.2026 02:00", 42.0])
    workbook.save(path)
    workbook.close()


def _write_template(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Comparatif"
    workbook.active.append(["reference"])
    workbook.save(path)
    workbook.close()


def test_captures_exact_bytes_gaps_and_disappeared_member(tmp_path: Path) -> None:
    _write_curve(tmp_path / "HFC_Ompex_20260101_101700.xlsx")
    _write_curve(tmp_path / "HFC_Ompex_20260103_101701.xlsx")
    _write_template(tmp_path / "Template_Comparatif.xlsx")

    audit = capture_ompex_archive_inventory(
        tmp_path,
        previous_observation={
            "file_count": 4,
            "member_set_complete": True,
            "file_names": [
                "HFC_Ompex_20260101_101700.xlsx",
                "HFC_Ompex_20260103_101701.xlsx",
                "Template_Comparatif.xlsx",
                "Template__HFC_Ompex_15min.xlsx",
            ],
        },
    )

    assert len(audit.files) == 3
    assert audit.summary["curve_vintage_count"] == 2
    assert audit.summary["missing_filename_dates"] == ["2026-01-02"]
    assert audit.summary["missing_filename_date_rate"] == pytest.approx(1 / 3)
    assert audit.summary["previous_observation"]["disappeared_file_names"] == [
        "Template__HFC_Ompex_15min.xlsx"
    ]
    assert audit.summary["origin_readiness"] == {
        "candidate_filename_origin_count": 2,
        "countable_scientific_origin_count": 0,
        "exact_vintage_hash_available": True,
        "availability_at_origin_authenticated": False,
        "rolling_origin_scoring_authorized": False,
    }
    assert audit.summary["price_values_or_statistics_emitted"] is False
    codes = {finding["code"] for finding in audit.summary["findings"]}
    assert "SOURCE_ARCHIVE_MEMBER_DISAPPEARED" in codes
    assert "FILENAME_DATE_COVERAGE_GAPS" in codes
    assert "FILENAME_AVAILABILITY_NOT_AUTHENTICATED" in codes


def test_same_archive_produces_same_price_free_inventory(tmp_path: Path) -> None:
    _write_curve(tmp_path / "HFC_Ompex_20260101_101700.xlsx")
    first = capture_ompex_archive_inventory(tmp_path)
    second = capture_ompex_archive_inventory(tmp_path)

    assert first == second
    row = first.files[0]
    assert len(str(row["sha256"])) == 64
    assert row["role"] == "curve_vintage"
    assert row["data_row_count"] == 2
    assert row["formula_tag_present"] is False
    assert row["price_values_or_statistics_emitted"] is False
    assert "price" not in row


def test_formula_curve_is_detected_and_fails_quality_status(tmp_path: Path) -> None:
    _write_curve(
        tmp_path / "HFC_Ompex_20260101_101700.xlsx", formula=True
    )
    audit = capture_ompex_archive_inventory(tmp_path)

    assert audit.summary["status"].startswith("FAIL_CURRENT_BYTE_INVENTORY")
    assert audit.summary["curve_formula_file_count"] == 1
    assert "CURVE_FORMULA_CONTENT_PRESENT" in {
        finding["code"] for finding in audit.summary["findings"]
    }


def test_rejects_wrong_curve_schema_and_non_xlsx_member(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "wrong"
    workbook.active.append(["Date", "EUR/MWh"])
    workbook.active.append(["01.01.2026 01:00", 1.0])
    workbook.save(tmp_path / "HFC_Ompex_20260101_101700.xlsx")
    workbook.close()
    with pytest.raises(OmpexArchiveInventoryError, match="unexpected sheets"):
        capture_ompex_archive_inventory(tmp_path)

    (tmp_path / "note.txt").write_text("not part of the archive", encoding="utf-8")
    with pytest.raises(OmpexArchiveInventoryError, match="non-XLSX"):
        capture_ompex_archive_inventory(tmp_path)


def test_rejects_duplicate_filename_dates(tmp_path: Path) -> None:
    _write_curve(tmp_path / "HFC_Ompex_20260101_101700.xlsx")
    _write_curve(tmp_path / "HFC_Ompex_20260101_101701.xlsx")

    with pytest.raises(OmpexArchiveInventoryError, match="duplicate filename"):
        capture_ompex_archive_inventory(tmp_path)


def test_recursively_inventories_regular_xlsx_members(tmp_path: Path) -> None:
    nested = tmp_path / "Template"
    nested.mkdir()
    _write_template(nested / "HFC_Ompex_15min.xlsx")
    _write_curve(tmp_path / "HFC_Ompex_20260101_101700.xlsx")

    audit = capture_ompex_archive_inventory(tmp_path)

    assert [row["file_name"] for row in audit.files] == [
        "HFC_Ompex_20260101_101700.xlsx",
        "Template/HFC_Ompex_15min.xlsx",
    ]
