from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from scripts.build_ch_hfc_validation_workbook import build_workbook


def test_build_validation_workbook_creates_expected_sheets(tmp_path):
    ts = pd.date_range("2030-01-01", periods=24 * 40, freq="h", tz="Europe/Zurich")
    rows = []
    for t in ts:
        price = 80.0
        if 11 <= t.hour <= 15:
            price -= 10.0
        if 17 <= t.hour <= 21:
            price += 12.0
        if t.weekday() >= 5:
            price -= 4.0
        rows.append(
            {
                "timestamp_ch": t.strftime("%d.%m.%Y %H:%M"),
                "utc_offset_ch": "UTC" + t.strftime("%z")[:3] + ":" + t.strftime("%z")[3:],
                "timestamp_utc": t.tz_convert("UTC").strftime("%d.%m.%Y %H:%M"),
                "price_slow_eur_mwh": price - 4.0,
                "price_central_eur_mwh": price,
                "price_fast_eur_mwh": price + 4.0,
                "price_weighted_mean_eur_mwh": price,
                "structural_p10_eur_mwh": price - 4.0,
                "structural_p50_eur_mwh": price,
                "structural_p90_eur_mwh": price + 4.0,
                "structural_width_eur_mwh": 8.0,
            }
        )
    csv = tmp_path / "hfc.csv"
    pd.DataFrame(rows).to_csv(csv, index=False)
    forwards = tmp_path / "forwards.parquet"
    pd.DataFrame(
        {
            "market": ["CH"],
            "load_type": ["BASE"],
            "date": [pd.Timestamp("2026-06-12")],
            "product": ["2030"],
            "price": [pd.read_csv(csv)["price_weighted_mean_eur_mwh"].mean()],
        }
    ).to_parquet(forwards, index=False)
    output = tmp_path / "validation.xlsx"
    report = tmp_path / "report.md"

    build_workbook(csv, forwards, output, report)

    assert output.exists()
    assert report.exists()
    wb = load_workbook(output, read_only=False)
    for sheet in [
        "README",
        "Raw",
        "EEX_Means",
        "Period_Means",
        "Duck_Month",
        "Duck_Season",
        "Heatmap_Month_Hour",
        "Structural_Width",
        "Negative_Low_Hours",
        "Seasonal_Coherence",
        "Quoted_EEX_Products",
        "Charts",
    ]:
        assert sheet in wb.sheetnames
    assert wb["Chart_Data"].sheet_state == "hidden"

    chart_output = tmp_path / "validation_charts_no_tables.xlsx"
    chart_report = tmp_path / "chart_report.md"
    build_workbook(
        csv,
        forwards,
        chart_output,
        chart_report,
        include_tables=False,
        hide_chart_data=False,
    )

    wb_charts = load_workbook(chart_output, read_only=False)
    assert len(wb_charts["Charts"]._charts) == 4
    assert sum(len(ws.tables) for ws in wb_charts.worksheets) == 0
    assert wb_charts["Chart_Data"].sheet_state == "visible"
